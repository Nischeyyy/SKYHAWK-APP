"""Skyhawk Security Operations - Backend API
JWT-based authentication, shift management, time clock, wallet, incidents, payroll, announcements.
"""
import io
import re
import os
import uuid
import json
import logging
import secrets
import mimetypes
from pathlib import Path
from datetime import datetime, timedelta, timezone, date
from typing import Optional, List, Annotated
from contextlib import asynccontextmanager

import jwt
import bcrypt
import httpx
from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Query, Request, UploadFile, File, Form
from fastapi.responses import JSONResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from pydantic import BaseModel, EmailStr, Field
from pymongo.errors import PyMongoError, ServerSelectionTimeoutError  # noqa: F401
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas as pdf_canvas
import notifications as notif

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

def _require_env(key: str) -> str:
    """Read a required environment variable; fail fast with a clear message if absent."""
    v = os.environ.get(key)
    if not v:
        raise RuntimeError(
            f"Required environment variable '{key}' is not set. "
            "Add it to Replit Secrets and restart."
        )
    return v


MONGO_URL = _require_env("MONGO_URL")
DB_NAME = _require_env("DB_NAME")
JWT_SECRET = _require_env("JWT_SECRET")
JWT_ALGORITHM = os.environ.get("JWT_ALGORITHM", "HS256")
JWT_EXPIRE_HOURS = int(os.environ.get("JWT_EXPIRE_HOURS", "168"))
EMERGENT_PUSH_KEY = os.environ.get("EMERGENT_PUSH_KEY", "placeholder")
PUSH_BASE_URL = "https://integrations.emergentagent.com"
# The external push provider requires a real key; without one, skip the network
# call entirely instead of generating a 401 for every notification attempt.
PUSH_ENABLED = bool(EMERGENT_PUSH_KEY) and EMERGENT_PUSH_KEY != "placeholder"

UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
(UPLOAD_DIR / "paystubs").mkdir(exist_ok=True)
(UPLOAD_DIR / "attachments").mkdir(exist_ok=True)
MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB
ALLOWED_UPLOAD_TYPES = {
    "image/jpeg", "image/png", "image/webp", "image/heic",
    "application/pdf",
}

RESET_TOKEN_TTL_MINUTES = 30

class _JSONFormatter(logging.Formatter):
    """Emit one JSON object per log line for easy machine parsing."""
    _SKIP = frozenset(logging.LogRecord.__dict__) | {
        "args", "msg", "message", "taskName",
    }

    def format(self, record: logging.LogRecord) -> str:
        record.message = record.getMessage()
        doc: dict = {
            "ts": self.formatTime(record, datefmt="%Y-%m-%dT%H:%M:%SZ"),
            "level": record.levelname,
            "logger": record.name,
            "msg": record.message,
        }
        if record.exc_info:
            doc["exception"] = self.formatException(record.exc_info)
        extra = {k: v for k, v in record.__dict__.items() if k not in self._SKIP}
        if extra:
            doc["extra"] = extra
        return json.dumps(doc)


_handler = logging.StreamHandler()
_handler.setFormatter(_JSONFormatter())
logging.root.handlers = [_handler]
logging.root.setLevel(logging.INFO)
logger = logging.getLogger("skyhawk")

client = AsyncIOMotorClient(
    MONGO_URL,
    serverSelectionTimeoutMS=5_000,   # fail fast if Atlas is unreachable
    connectTimeoutMS=10_000,
    socketTimeoutMS=30_000,
    maxPoolSize=50,
    minPoolSize=5,
    retryWrites=True,
)
db = client[DB_NAME]

# Rate limiter — shared across all routes
limiter = Limiter(key_func=get_remote_address, default_limits=["300/minute"])

push_client = httpx.AsyncClient(
    base_url=PUSH_BASE_URL,
    headers={"X-Push-Key": EMERGENT_PUSH_KEY},
    timeout=10.0,
)


# ============================================================
# Utilities
# ============================================================
def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def iso(dt: datetime) -> str:
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.isoformat()


def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False


def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "role": role,
        "exp": now_utc() + timedelta(hours=JWT_EXPIRE_HOURS),
        "iat": now_utc(),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


bearer = HTTPBearer(auto_error=False)


async def get_current_user(creds: Optional[HTTPAuthorizationCredentials] = Depends(bearer)):
    if not creds:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Not authenticated")
    try:
        payload = jwt.decode(creds.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Invalid token")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Token expired")
    except jwt.PyJWTError:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Invalid token")

    user = await db.users.find_one({"id": user_id}, {"_id": 0, "hashed_password": 0})
    if not user:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "User not found")
    return user


async def send_push(recipients: list[str], data: dict, idempotency_key: Optional[str] = None):
    if not recipients:
        return
    if "title" not in data or "message" not in data:
        return
    if not PUSH_ENABLED:
        # No real push provider key configured — this is expected in dev/demo
        # environments. Avoid spamming logs with a 401 per notification.
        return
    payload = {"recipients": recipients, "data": data}
    if idempotency_key:
        payload["$idempotency_key"] = idempotency_key
    try:
        resp = await push_client.post("/api/v1/push/trigger", json=payload)
        if resp.status_code >= 400:
            logger.warning(f"Push failed: {resp.status_code} {resp.text}")
    except Exception as e:
        logger.warning(f"Push notification failed (non-blocking): {e}")


# ============================================================
# Pydantic Models
# ============================================================
class RegisterIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    full_name: str
    phone: Optional[str] = None


class LoginIn(BaseModel):
    email: EmailStr
    password: str


class AuthOut(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict


class ClockInIn(BaseModel):
    shift_id: Optional[str] = None
    latitude: float
    longitude: float
    selfie_base64: str
    site_id: Optional[str] = None


class ClockOutIn(BaseModel):
    latitude: float
    longitude: float
    selfie_base64: Optional[str] = None


class BreakIn(BaseModel):
    action: str  # "start" | "end"


class AckIn(BaseModel):
    pass


_Photo = Annotated[str, Field(max_length=500_000)]  # ~375 KB per image

class IncidentIn(BaseModel):
    type: str = Field(max_length=50)  # incident|injury|lost_found|property_damage
    site_id: Optional[str] = Field(default=None, max_length=100)
    description: str = Field(min_length=5, max_length=5_000)
    severity: str = Field(default="medium", max_length=20)  # low|medium|high|critical
    witness_name: Optional[str] = Field(default=None, max_length=200)
    witness_contact: Optional[str] = Field(default=None, max_length=200)
    photos: List[_Photo] = Field(default=[], max_length=5)  # max 5 base64 images
    signature_base64: Optional[str] = Field(default=None, max_length=500_000)


class RegisterPushIn(BaseModel):
    user_id: str
    platform: str
    device_token: str


class ClaimShiftIn(BaseModel):
    pass


class SiteAckIn(BaseModel):
    site_id: str


class SOSIn(BaseModel):
    latitude: float
    longitude: float
    message: Optional[str] = Field(default=None, max_length=500)


class IncidentStatusIn(BaseModel):
    status: str = Field(max_length=30)
    note: Optional[str] = Field(default=None, max_length=2000)
    assigned_to: Optional[str] = Field(default=None, max_length=100)


class LocationPingIn(BaseModel):
    latitude: float
    longitude: float
    accuracy_m: Optional[float] = None


class SwapRequestIn(BaseModel):
    reason: Optional[str] = Field(default=None, max_length=500)


class SwapActionIn(BaseModel):
    action: str  # approve | reject


class PayrollCalculateIn(BaseModel):
    user_id: str
    period_start: str
    period_end: str
    pay_date: str
    hourly_rate: float = Field(gt=0)
    overtime_threshold: float = 40.0
    overtime_multiplier: float = 1.5
    tax_rate: float = 0.28


class WalletDocIn(BaseModel):
    type: str = Field(max_length=50)
    name: str = Field(max_length=200)
    number: str = Field(max_length=100)
    expiry: str = Field(max_length=50)


class CommunityPostIn(BaseModel):
    body: str = Field(min_length=1, max_length=4000)
    type: str = Field(default="post", max_length=20)  # post | announcement | event | recognition
    audience: str = Field(default="All Staff", max_length=50)
    attachments: List[dict] = Field(default_factory=list)


class CommunityCommentIn(BaseModel):
    body: str = Field(min_length=1, max_length=1000)


class CommunityEditIn(BaseModel):
    body: str = Field(min_length=1, max_length=4000)
    attachments: List[dict] = Field(default_factory=list)


class ChangePasswordIn(BaseModel):
    current_password: str
    new_password: str = Field(min_length=6)


class ForgotPasswordIn(BaseModel):
    email: EmailStr


class ResetPasswordIn(BaseModel):
    token: str
    new_password: str = Field(min_length=6)


class SettingsIn(BaseModel):
    push_notifications: Optional[bool] = None
    location_sharing: Optional[bool] = None
    shift_reminders: Optional[bool] = None
    community_notifications: Optional[bool] = None


class OnboardingDocIn(BaseModel):
    type: str = Field(max_length=50)
    name: str = Field(max_length=200)
    number: str = Field(max_length=100)
    expiry: str = Field(max_length=50)
    attachment_url: Optional[str] = Field(default=None, max_length=500)


class OnboardingSINIn(BaseModel):
    sin: str = Field(min_length=9, max_length=11)


class OnboardingDirectDepositIn(BaseModel):
    institution: str = Field(max_length=10)
    transit: str = Field(max_length=10)
    account: str = Field(max_length=20)


class OnboardingEmergencyContactIn(BaseModel):
    name: str = Field(max_length=200)
    phone: str = Field(max_length=50)
    relation: str = Field(max_length=50)


class OnboardingAgreementsIn(BaseModel):
    signed: bool


# ============================================================
# App
# ============================================================
@asynccontextmanager
async def lifespan(app: FastAPI):
    try:
        await ensure_indexes()
        logger.info("Database indexes ensured")
    except Exception as exc:
        logger.error("Failed to create DB indexes — check MONGO_URL: %s", exc)
    notif.log_status()
    try:
        await seed_data()
    except Exception as exc:
        logger.error("Failed to seed demo data: %s", exc)
    yield
    client.close()
    await push_client.aclose()


app = FastAPI(title="Skyhawk Ops API", lifespan=lifespan)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)


@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception) -> JSONResponse:
    """Catch-all for unhandled exceptions — log with context, return a safe 500."""
    if isinstance(exc, HTTPException):
        raise exc
    logger.error(
        "Unhandled exception on %s %s",
        request.method,
        request.url.path,
        exc_info=exc,
    )
    return JSONResponse(
        status_code=500,
        content={"detail": "An unexpected error occurred. Please try again."},
    )


api = APIRouter(prefix="/api")


@api.get("/health")
async def health():
    """Liveness + readiness probe. Returns 503 if MongoDB is unreachable."""
    try:
        await db.command("ping")
        return {"status": "ok", "database": "connected"}
    except Exception as exc:
        logger.warning("Health check failed: %s", exc)
        raise HTTPException(status_code=503, detail="Database unavailable")


async def ensure_indexes():
    await db.users.create_index("email", unique=True)
    await db.users.create_index("id", unique=True)
    await db.shifts.create_index("id", unique=True)
    await db.open_shifts.create_index("id", unique=True)
    await db.sites.create_index("id", unique=True)
    await db.announcements.create_index("id", unique=True)
    await db.incidents.create_index("id", unique=True)
    await db.timeclock.create_index("id", unique=True)
    await db.payroll.create_index("id", unique=True)
    await db.push_tokens.create_index([("user_id", 1), ("device_token", 1)], unique=True)
    await db.sos_alerts.create_index("id", unique=True)
    await db.sos_alerts.create_index([("status", 1), ("created_at", -1)])
    await db.shift_swaps.create_index("id", unique=True)
    await db.shift_swaps.create_index([("status", 1), ("start", 1)])
    await db.location_pings.create_index([("timeclock_id", 1), ("ts", -1)])
    await db.community_posts.create_index("id", unique=True)
    await db.community_posts.create_index([("type", 1), ("created_at", -1)])
    await db.password_resets.create_index("user_id", unique=True)
    await db.notifications.create_index([("user_id", 1), ("created_at", -1)])
    await db.notifications.create_index([("user_id", 1), ("read", 1)])


# ============================================================
# Seed demo data
# ============================================================
async def seed_data():
    # Idempotent seed
    if await db.users.count_documents({}) > 0:
        return
    logger.info("Seeding demo data...")

    admin_id = str(uuid.uuid4())
    guard_id = str(uuid.uuid4())
    guard2_id = str(uuid.uuid4())

    await db.users.insert_many([
        {
            "id": admin_id,
            "email": "admin@skyhawk.com",
            "hashed_password": hash_password("Admin123"),
            "full_name": "Admin Steele",
            "phone": "+1 416 555 0000",
            "role": "admin",
            "employee_number": "SH-0001",
            "licence_number": "SG-ADM-2026",
            "licence_expiry": iso(now_utc() + timedelta(days=400)),
            "certifications": ["First Aid", "Smart Serve", "WHMIS", "CPR"],
            "employment_status": "Active - Full Time",
            "photo_url": "https://images.unsplash.com/photo-1547882230-87a3d4390e8d",
            "emergency_contact": {"name": "Sara Steele", "phone": "+1 416 555 0111", "relation": "Spouse"},
            "onboarding_complete": True,
            "created_at": iso(now_utc()),
        },
        {
            "id": guard_id,
            "email": "guard@skyhawk.com",
            "hashed_password": hash_password("Password123"),
            "full_name": "Marcus Vance",
            "phone": "+1 416 555 0142",
            "role": "employee",
            "employee_number": "SH-2041",
            "licence_number": "SG-ON-89442",
            "licence_expiry": iso(now_utc() + timedelta(days=42)),
            "certifications": ["First Aid", "WHMIS", "Smart Serve"],
            "employment_status": "Active - Full Time",
            "photo_url": "https://images.unsplash.com/photo-1547882230-87a3d4390e8d",
            "emergency_contact": {"name": "Elena Vance", "phone": "+1 416 555 0999", "relation": "Sister"},
            "onboarding_complete": True,
            "created_at": iso(now_utc()),
        },
        {
            "id": guard2_id,
            "email": "guard2@skyhawk.com",
            "hashed_password": hash_password("Password123"),
            "full_name": "Priya Kaur",
            "phone": "+1 647 555 0212",
            "role": "employee",
            "employee_number": "SH-2088",
            "licence_number": "SG-ON-91201",
            "licence_expiry": iso(now_utc() + timedelta(days=180)),
            "certifications": ["First Aid", "CPR"],
            "employment_status": "Active - Part Time",
            "photo_url": "https://images.unsplash.com/photo-1544005313-94ddf0286df2",
            "emergency_contact": {"name": "Ravi Kaur", "phone": "+1 647 555 0111", "relation": "Father"},
            "onboarding_complete": False,
            "created_at": iso(now_utc()),
        },
    ])

    # Sites
    site_ids = [str(uuid.uuid4()) for _ in range(4)]
    sites = [
        {
            "id": site_ids[0], "name": "Toronto Financial Tower",
            "address": "88 Bay Street, Toronto, ON M5J 2S1",
            "latitude": 43.6467, "longitude": -79.3785,
            "supervisor": "Derek Cole", "supervisor_phone": "+1 416 555 0301",
            "instructions": "Report to lobby desk. Uniform inspection required. Rounds every 30 minutes. Log all entries in the electronic register. No mobile use on patrol.",
            "geofence_radius_m": 150,
        },
        {
            "id": site_ids[1], "name": "Skyline Convention Centre",
            "address": "255 Front Street West, Toronto, ON M5V 2W6",
            "latitude": 43.6427, "longitude": -79.3860,
            "supervisor": "Alicia Ng", "supervisor_phone": "+1 416 555 0342",
            "instructions": "Enter via service entrance C. Check credentials at all main doors. Escalate any suspicious packages immediately to supervisor.",
            "geofence_radius_m": 200,
        },
        {
            "id": site_ids[2], "name": "Northlake Warehouse",
            "address": "1200 Pharmacy Ave, Scarborough, ON M1P 2M3",
            "latitude": 43.7691, "longitude": -79.2966,
            "supervisor": "Jamal Reeves", "supervisor_phone": "+1 416 555 0555",
            "instructions": "K9 unit on-site nights. Loading bay 4 requires escort access. Motion sensors monitored — reset via panel in office 2B.",
            "geofence_radius_m": 250,
        },
        {
            "id": site_ids[3], "name": "Harbourfront Residences",
            "address": "10 Queens Quay West, Toronto, ON M5J 2R9",
            "latitude": 43.6389, "longitude": -79.3806,
            "supervisor": "Fiona Marsh", "supervisor_phone": "+1 416 555 0212",
            "instructions": "Concierge relief 22:00 - 06:00. Verify all guest passes. Escalate noise complaints to property manager.",
            "geofence_radius_m": 120,
        },
    ]
    await db.sites.insert_many(sites)

    # Today's active shift for guard
    today = now_utc().replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = today + timedelta(days=1)
    shift_start = today.replace(hour=18, minute=0)
    shift_end = today.replace(hour=23, minute=59) + timedelta(hours=2)  # ends 02:00 next day

    scheduled_shifts = []
    # Today's shift (active/upcoming)
    scheduled_shifts.append({
        "id": str(uuid.uuid4()), "user_id": guard_id, "site_id": site_ids[0],
        "start": iso(shift_start), "end": iso(shift_end),
        "role": "Concierge Security", "pay_rate": 24.50, "status": "scheduled",
        "instructions_acknowledged": False,
    })
    # Tomorrow
    scheduled_shifts.append({
        "id": str(uuid.uuid4()), "user_id": guard_id, "site_id": site_ids[1],
        "start": iso(tomorrow.replace(hour=22)), "end": iso(tomorrow.replace(hour=23, minute=59) + timedelta(hours=6, minutes=1)),
        "role": "Event Security", "pay_rate": 26.00, "status": "scheduled",
        "instructions_acknowledged": True,
    })
    # This week upcoming
    for i in range(2, 6):
        d = today + timedelta(days=i)
        scheduled_shifts.append({
            "id": str(uuid.uuid4()), "user_id": guard_id, "site_id": site_ids[i % 4],
            "start": iso(d.replace(hour=18)),
            "end": iso(d.replace(hour=23, minute=59) + timedelta(hours=2, minutes=1)),
            "role": "Patrol", "pay_rate": 24.50, "status": "scheduled",
            "instructions_acknowledged": False,
        })
    # Past week completed
    for i in range(1, 6):
        d = today - timedelta(days=i)
        scheduled_shifts.append({
            "id": str(uuid.uuid4()), "user_id": guard_id, "site_id": site_ids[i % 4],
            "start": iso(d.replace(hour=18)),
            "end": iso(d.replace(hour=23, minute=59) + timedelta(hours=2, minutes=1)),
            "role": "Patrol", "pay_rate": 24.50, "status": "completed",
            "hours_worked": 8.0, "instructions_acknowledged": True,
        })
    await db.shifts.insert_many(scheduled_shifts)

    # Open shifts marketplace
    open_shifts = []
    for i, offset in enumerate([1, 2, 3, 4, 5, 6, 7]):
        d = today + timedelta(days=offset)
        open_shifts.append({
            "id": str(uuid.uuid4()),
            "site_id": site_ids[i % 4],
            "start": iso(d.replace(hour=18 if i % 2 == 0 else 8)),
            "end": iso(d.replace(hour=23, minute=59) + timedelta(hours=2 if i % 2 == 0 else -7)),
            "role": ["Patrol", "Concierge", "Event Security", "K9 Support"][i % 4],
            "pay_rate": [24.5, 26.0, 28.0, 30.0][i % 4],
            "spots_available": (i % 3) + 1,
            "posted_at": iso(now_utc() - timedelta(hours=i * 3)),
            "claimed_by": [],
            "waitlist": [],
            "urgent": i < 2,
        })
    await db.open_shifts.insert_many(open_shifts)

    # Announcements
    announcements = [
        {
            "id": str(uuid.uuid4()), "title": "Winter Uniform Rollout",
            "body": "All guards must collect winter jackets from HQ by Nov 15. Sign the equipment log on pickup.",
            "severity": "info", "site_id": None, "posted_by": "Operations",
            "posted_at": iso(now_utc() - timedelta(hours=4)),
            "read_by": [],
        },
        {
            "id": str(uuid.uuid4()), "title": "URGENT: Elevator Outage - Bay Street",
            "body": "Elevator 3 at 88 Bay St is out of service until 22:00. Use stairwell C for patrol rounds. Report anyone using it.",
            "severity": "critical", "site_id": site_ids[0], "posted_by": "Derek Cole",
            "posted_at": iso(now_utc() - timedelta(hours=1)),
            "read_by": [],
        },
        {
            "id": str(uuid.uuid4()), "title": "Payroll Cycle Update",
            "body": "This period closes Sunday 23:59. Ensure all timecards are submitted with GPS + selfie verification.",
            "severity": "info", "site_id": None, "posted_by": "Payroll",
            "posted_at": iso(now_utc() - timedelta(days=1)),
            "read_by": [guard_id],
        },
        {
            "id": str(uuid.uuid4()), "title": "Emergency Contact Drill",
            "body": "Full drill Thu 03:00. All patrol staff must acknowledge this notice by 20:00.",
            "severity": "warning", "site_id": None, "posted_by": "HQ Command",
            "posted_at": iso(now_utc() - timedelta(hours=8)),
            "read_by": [],
        },
    ]
    await db.announcements.insert_many(announcements)

    # Payroll periods
    payroll_periods = []
    for i in range(4):
        period_end = today - timedelta(days=i * 14)
        period_start = period_end - timedelta(days=13)
        stages = ["paid", "released", "under_review", "submitted"]
        payroll_periods.append({
            "id": str(uuid.uuid4()),
            "user_id": guard_id,
            "period_start": iso(period_start),
            "period_end": iso(period_end),
            "pay_date": iso(period_end + timedelta(days=5)),
            "hours_regular": 76.5 - i * 2,
            "hours_overtime": 4.0 if i == 0 else 0.0,
            "gross": (76.5 - i * 2) * 24.5 + (98 if i == 0 else 0),
            "net": ((76.5 - i * 2) * 24.5 + (98 if i == 0 else 0)) * 0.72,
            "status": stages[i],
            "pay_stub_url": None,
        })
    await db.payroll.insert_many(payroll_periods)

    # Wallet documents
    wallet_docs = [
        {"id": str(uuid.uuid4()), "user_id": guard_id, "type": "security_licence",
         "name": "Ontario Security Guard Licence", "number": "SG-ON-89442",
         "expiry": iso(now_utc() + timedelta(days=42)), "status": "expiring_soon"},
        {"id": str(uuid.uuid4()), "user_id": guard_id, "type": "company_id",
         "name": "Skyhawk Company ID", "number": "SH-2041",
         "expiry": iso(now_utc() + timedelta(days=800)), "status": "valid"},
        {"id": str(uuid.uuid4()), "user_id": guard_id, "type": "first_aid",
         "name": "First Aid & CPR", "number": "FA-2024-88112",
         "expiry": iso(now_utc() + timedelta(days=210)), "status": "valid"},
        {"id": str(uuid.uuid4()), "user_id": guard_id, "type": "smart_serve",
         "name": "Smart Serve Ontario", "number": "SS-2025-4421",
         "expiry": iso(now_utc() + timedelta(days=500)), "status": "valid"},
        {"id": str(uuid.uuid4()), "user_id": guard_id, "type": "whmis",
         "name": "WHMIS 2015", "number": "WH-2025-9012",
         "expiry": iso(now_utc() + timedelta(days=300)), "status": "valid"},
        {"id": str(uuid.uuid4()), "user_id": guard_id, "type": "work_permit",
         "name": "Work Permit", "number": "WP-CA-88214",
         "expiry": iso(now_utc() + timedelta(days=900)), "status": "valid"},
    ]
    await db.wallet_documents.insert_many(wallet_docs)

    # Equipment assigned
    await db.equipment.insert_many([
        {"id": str(uuid.uuid4()), "user_id": guard_id, "type": "uniform",
         "description": "Winter Jacket XL / Trousers 34", "issued_at": iso(now_utc() - timedelta(days=180)), "status": "issued"},
        {"id": str(uuid.uuid4()), "user_id": guard_id, "type": "radio",
         "description": "Motorola CP200d #R-4412", "issued_at": iso(now_utc() - timedelta(days=90)), "status": "issued"},
        {"id": str(uuid.uuid4()), "user_id": guard_id, "type": "keys",
         "description": "Master keyring #K-88 (Bay St)", "issued_at": iso(now_utc() - timedelta(days=60)), "status": "issued"},
    ])

    logger.info("Seed complete.")


# ============================================================
# AUTH
# ============================================================
@api.get("/")
async def root():
    return {"service": "Skyhawk Ops API", "status": "online"}


@api.post("/auth/register", response_model=AuthOut)
@limiter.limit("5/minute")
async def register(request: Request, body: RegisterIn):
    existing = await db.users.find_one({"email": body.email.lower()})
    if existing:
        raise HTTPException(400, "Email already registered")
    user_id = str(uuid.uuid4())
    user = {
        "id": user_id,
        "email": body.email.lower(),
        "hashed_password": hash_password(body.password),
        "full_name": body.full_name,
        "phone": body.phone,
        "role": "employee",
        "employee_number": f"SH-{str(uuid.uuid4())[:4].upper()}",
        "licence_number": None,
        "licence_expiry": None,
        "certifications": [],
        "employment_status": "Onboarding",
        "photo_url": None,
        "emergency_contact": None,
        "onboarding_complete": False,
        "preferences": {
            "push_notifications": True,
            "location_sharing": True,
            "shift_reminders": True,
            "community_notifications": True,
        },
        "created_at": iso(now_utc()),
    }
    await db.users.insert_one(user)
    token = create_access_token(user_id, user["email"], "employee")
    user.pop("hashed_password", None)
    user.pop("_id", None)
    return {"access_token": token, "token_type": "bearer", "user": user}


@api.post("/auth/login", response_model=AuthOut)
@limiter.limit("10/minute")
async def login(request: Request, body: LoginIn):
    user = await db.users.find_one({"email": body.email.lower()})
    if not user or not verify_password(body.password, user["hashed_password"]):
        raise HTTPException(401, "Invalid email or password")
    token = create_access_token(user["id"], user["email"], user["role"])
    user.pop("hashed_password", None)
    user.pop("_id", None)
    return {"access_token": token, "token_type": "bearer", "user": user}


@api.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return user


@api.post("/auth/change-password")
@limiter.limit("5/minute")
async def change_password(request: Request, body: ChangePasswordIn, user=Depends(get_current_user)):
    full_user = await db.users.find_one({"id": user["id"]})
    if not full_user or not verify_password(body.current_password, full_user["hashed_password"]):
        raise HTTPException(401, "Current password is incorrect")
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"hashed_password": hash_password(body.new_password)}},
    )
    # Invalidate any outstanding reset tokens once the password changes.
    await db.password_resets.delete_many({"user_id": user["id"]})
    return {"status": "password_updated"}


@api.post("/auth/forgot-password")
@limiter.limit("5/minute")
async def forgot_password(request: Request, body: ForgotPasswordIn):
    user = await db.users.find_one({"email": body.email.lower()})
    # Always return a generic response so this endpoint can't be used to
    # enumerate registered emails.
    generic = {"status": "if_registered_email_sent"}
    if not user:
        return generic
    raw_token = secrets.token_urlsafe(32)
    await db.password_resets.update_one(
        {"user_id": user["id"]},
        {"$set": {
            "user_id": user["id"],
            "token_hash": hash_password(raw_token),
            "expires_at": iso(now_utc() + timedelta(minutes=RESET_TOKEN_TTL_MINUTES)),
            "created_at": iso(now_utc()),
        }},
        upsert=True,
    )
    reset_link = f"skyhawksecurity://reset-password?token={raw_token}"
    html = notif.email_template(
        "Reset your Skyhawk password",
        [
            f"Hi {user.get('full_name', 'there')},",
            "We received a request to reset your password. Click the button below — "
            "the link expires in <strong>30 minutes</strong>.",
            "If you didn't request this, you can safely ignore this email.",
        ],
        cta_text="Reset Password",
        cta_link=reset_link,
    )
    sent = await notif.send_email(
        user["email"],
        "Reset your Skyhawk Security password",
        html,
        f"Password reset link (expires in 30 min): {reset_link}",
    )
    if not sent:
        logger.info(
            "Password reset token (email not configured)",
            extra={"reset_token": raw_token, "user": user["email"]},
        )
    return generic


@api.post("/auth/reset-password")
@limiter.limit("5/minute")
async def reset_password(request: Request, body: ResetPasswordIn):
    candidates = await db.password_resets.find({"expires_at": {"$gte": iso(now_utc())}}).to_list(1000)
    match = next((c for c in candidates if verify_password(body.token, c["token_hash"])), None)
    if not match:
        raise HTTPException(400, "Invalid or expired reset token")
    await db.users.update_one(
        {"id": match["user_id"]},
        {"$set": {"hashed_password": hash_password(body.new_password)}},
    )
    await db.password_resets.delete_many({"user_id": match["user_id"]})
    return {"status": "password_reset"}


# ============================================================
# SETTINGS / PREFERENCES
# ============================================================
@api.get("/settings")
async def get_settings(user=Depends(get_current_user)):
    defaults = {
        "push_notifications": True,
        "location_sharing": True,
        "shift_reminders": True,
        "community_notifications": True,
    }
    return {**defaults, **(user.get("preferences") or {})}


@api.put("/settings")
async def update_settings(body: SettingsIn, user=Depends(get_current_user)):
    updates = {f"preferences.{k}": v for k, v in body.model_dump(exclude_none=True).items()}
    if updates:
        await db.users.update_one({"id": user["id"]}, {"$set": updates})
    fresh = await db.users.find_one({"id": user["id"]}, {"_id": 0, "hashed_password": 0})
    defaults = {
        "push_notifications": True,
        "location_sharing": True,
        "shift_reminders": True,
        "community_notifications": True,
    }
    return {**defaults, **(fresh.get("preferences") or {})}


# ============================================================
# IN-APP NOTIFICATIONS
# ============================================================
@api.get("/notifications")
async def get_notifications(user=Depends(get_current_user), limit: int = 60):
    items = await db.notifications.find(
        {"user_id": user["id"]}, {"_id": 0}
    ).sort("created_at", -1).to_list(min(limit, 100))
    return {"notifications": items}


@api.get("/notifications/unread-count")
async def get_unread_count(user=Depends(get_current_user)):
    count = await db.notifications.count_documents({"user_id": user["id"], "read": False})
    return {"count": count}


@api.post("/notifications/read-all")
async def mark_all_notifications_read(user=Depends(get_current_user)):
    await db.notifications.update_many(
        {"user_id": user["id"], "read": False},
        {"$set": {"read": True}},
    )
    return {"status": "ok"}


@api.post("/notifications/{notif_id}/read")
async def mark_notification_read(notif_id: str, user=Depends(get_current_user)):
    await db.notifications.update_one(
        {"id": notif_id, "user_id": user["id"]},
        {"$set": {"read": True}},
    )
    return {"status": "ok"}


# ============================================================
# FILE UPLOADS
# ============================================================
@api.post("/uploads", status_code=201)
@limiter.limit("20/minute")
async def upload_file(request: Request, file: UploadFile = File(...), user=Depends(get_current_user)):
    if file.content_type not in ALLOWED_UPLOAD_TYPES:
        raise HTTPException(400, f"Unsupported file type: {file.content_type}")
    contents = await file.read()
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, "File too large (max 10 MB)")
    ext = mimetypes.guess_extension(file.content_type) or ""
    stored_name = f"{uuid.uuid4()}{ext}"
    dest = UPLOAD_DIR / "attachments" / stored_name
    dest.write_bytes(contents)
    return {
        "url": f"/uploads/attachments/{stored_name}",
        "name": file.filename,
        "content_type": file.content_type,
        "size": len(contents),
        "kind": "image" if file.content_type.startswith("image/") else "pdf",
    }


# ============================================================
# DASHBOARD
# ============================================================
@api.get("/dashboard")
@limiter.limit("60/minute")
async def dashboard(request: Request, user=Depends(get_current_user)):
    now = now_utc()
    uid = user["id"]

    # Today's shift
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)
    today_shift = await db.shifts.find_one(
        {"user_id": uid, "start": {"$gte": iso(today_start), "$lt": iso(today_end)}},
        {"_id": 0},
    )
    if today_shift:
        site = await db.sites.find_one({"id": today_shift["site_id"]}, {"_id": 0})
        today_shift["site"] = site

    # Next shift
    next_shift = await db.shifts.find_one(
        {"user_id": uid, "start": {"$gte": iso(now)}, "status": "scheduled"},
        {"_id": 0},
        sort=[("start", 1)],
    )
    if next_shift:
        site = await db.sites.find_one({"id": next_shift["site_id"]}, {"_id": 0})
        next_shift["site"] = site

    # Active clock-in
    active_clock = await db.timeclock.find_one(
        {"user_id": uid, "clock_out": None}, {"_id": 0, "selfie_in": 0, "selfie_out": 0}
    )

    # Announcements (unread)
    unread_ann = await db.announcements.count_documents({"read_by": {"$ne": uid}})

    # Latest payroll status
    latest_payroll = await db.payroll.find_one(
        {"user_id": uid}, {"_id": 0}, sort=[("period_end", -1)]
    )

    # Licence expiry
    licence_days_remaining = None
    if user.get("licence_expiry"):
        expiry = datetime.fromisoformat(user["licence_expiry"].replace("Z", "+00:00"))
        licence_days_remaining = (expiry - now).days

    return {
        "today_shift": today_shift,
        "next_shift": next_shift,
        "active_clock": active_clock,
        "unread_announcements": unread_ann,
        "latest_payroll": latest_payroll,
        "licence_days_remaining": licence_days_remaining,
        "emergency_contacts": [
            {"name": "HQ Dispatch", "phone": "+1 416 555 0000"},
            {"name": "911 Emergency", "phone": "911"},
            {"name": "Site Supervisor", "phone": (today_shift or {}).get("site", {}).get("supervisor_phone")},
        ],
    }


# ============================================================
# SCHEDULE
# ============================================================
@api.get("/schedule")
async def schedule(
    user=Depends(get_current_user),
    range: str = Query("week"),  # week|month
    start: Optional[str] = None,
    end: Optional[str] = None,
):
    now = now_utc()
    if start and end:
        start = datetime.fromisoformat(start)
        end = datetime.fromisoformat(end)
        if start.tzinfo is None:
            start = start.replace(tzinfo=timezone.utc)
        if end.tzinfo is None:
            end = end.replace(tzinfo=timezone.utc)
    elif range == "month":
        start = now - timedelta(days=15)
        end = now + timedelta(days=30)
    else:
        weekday = now.weekday()
        start = (now - timedelta(days=weekday)).replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=7)
    shifts = await db.shifts.find(
        {"user_id": user["id"], "start": {"$gte": iso(start), "$lt": iso(end)}},
        {"_id": 0},
    ).sort("start", 1).to_list(200)
    site_ids = list({s["site_id"] for s in shifts})
    sites = {s["id"]: s for s in await db.sites.find({"id": {"$in": site_ids}}, {"_id": 0}).to_list(50)}
    for s in shifts:
        s["site"] = sites.get(s["site_id"])
    return {"shifts": shifts, "range_start": iso(start), "range_end": iso(end)}


@api.get("/schedule/stats")
async def schedule_stats(user=Depends(get_current_user)):
    """Quick summary card: week hours, week earnings, upcoming count, pending acks, active clock."""
    now = now_utc()
    # Week boundaries (Monday start)
    weekday = now.weekday()
    week_start = (now - timedelta(days=weekday)).replace(hour=0, minute=0, second=0, microsecond=0)
    week_end = week_start + timedelta(days=7)

    week_shifts = await db.shifts.find(
        {"user_id": user["id"], "start": {"$gte": iso(week_start), "$lt": iso(week_end)}},
        {"_id": 0, "start": 1, "end": 1, "pay_rate": 1, "status": 1, "hours_worked": 1},
    ).to_list(100)

    week_hours = 0.0
    week_earnings = 0.0
    for s in week_shifts:
        if s.get("hours_worked") is not None:
            hrs = s["hours_worked"]
        elif s.get("start") and s.get("end"):
            try:
                hrs = (
                    datetime.fromisoformat(s["end"].replace("Z", "+00:00")) -
                    datetime.fromisoformat(s["start"].replace("Z", "+00:00"))
                ).total_seconds() / 3600
            except Exception:
                hrs = 0.0
        else:
            hrs = 0.0
        week_hours += hrs
        week_earnings += hrs * s.get("pay_rate", 0)

    upcoming = await db.shifts.count_documents({
        "user_id": user["id"], "status": "scheduled", "start": {"$gte": iso(now)},
    })
    needs_ack = await db.shifts.count_documents({
        "user_id": user["id"], "status": "scheduled",
        "instructions_acknowledged": False, "start": {"$gte": iso(now)},
    })
    active_clock = await db.timeclock.find_one(
        {"user_id": user["id"], "clock_out": None},
        {"_id": 0, "selfie_in": 0, "selfie_out": 0},
    )
    return {
        "week_hours": round(week_hours, 1),
        "week_earnings": round(week_earnings, 2),
        "upcoming_shifts": upcoming,
        "needs_ack": needs_ack,
        "active_clock": bool(active_clock),
    }


@api.get("/shifts/{shift_id}")
async def get_shift(shift_id: str, user=Depends(get_current_user)):
    s = await db.shifts.find_one({"id": shift_id, "user_id": user["id"]}, {"_id": 0})
    if not s:
        raise HTTPException(404, "Shift not found")
    s["site"] = await db.sites.find_one({"id": s["site_id"]}, {"_id": 0})
    return s


@api.post("/shifts/{shift_id}/acknowledge-instructions")
async def acknowledge_instructions(shift_id: str, user=Depends(get_current_user)):
    r = await db.shifts.update_one(
        {"id": shift_id, "user_id": user["id"]},
        {"$set": {"instructions_acknowledged": True, "instructions_acknowledged_at": iso(now_utc())}},
    )
    if r.matched_count == 0:
        raise HTTPException(404, "Shift not found")
    return {"acknowledged": True}


# ============================================================
# OPEN SHIFTS
# ============================================================
@api.get("/open-shifts")
async def list_open_shifts(
    user=Depends(get_current_user),
    site_id: Optional[str] = None,
    from_date: Optional[str] = None,
):
    q = {}
    if site_id:
        q["site_id"] = site_id
    if from_date:
        q["start"] = {"$gte": from_date}
    shifts = await db.open_shifts.find(q, {"_id": 0}).sort("start", 1).to_list(100)
    site_ids = list({s["site_id"] for s in shifts})
    sites = {s["id"]: s for s in await db.sites.find({"id": {"$in": site_ids}}, {"_id": 0}).to_list(50)}
    for s in shifts:
        s["site"] = sites.get(s["site_id"])
        s["already_claimed"] = user["id"] in (s.get("claimed_by") or [])
        s["on_waitlist"] = user["id"] in (s.get("waitlist") or [])
    return {"shifts": shifts}


@api.post("/open-shifts/{shift_id}/claim")
async def claim_shift(shift_id: str, user=Depends(get_current_user)):
    s = await db.open_shifts.find_one({"id": shift_id})
    if not s:
        raise HTTPException(404, "Shift not found")
    claimed = s.get("claimed_by") or []
    if user["id"] in claimed:
        raise HTTPException(400, "Already claimed")
    if len(claimed) >= s["spots_available"]:
        # Add to waitlist
        await db.open_shifts.update_one({"id": shift_id}, {"$addToSet": {"waitlist": user["id"]}})
        return {"status": "waitlisted"}
    await db.open_shifts.update_one({"id": shift_id}, {"$addToSet": {"claimed_by": user["id"]}})
    # Create a scheduled shift for the user
    new_shift = {
        "id": str(uuid.uuid4()), "user_id": user["id"], "site_id": s["site_id"],
        "start": s["start"], "end": s["end"], "role": s["role"],
        "pay_rate": s["pay_rate"], "status": "scheduled",
        "claimed_from_open": shift_id,
        "instructions_acknowledged": False,
    }
    await db.shifts.insert_one(new_shift)
    await send_push([user["id"]], {"title": "Shift Claimed", "message": f"You claimed {s['role']} on {s['start'][:10]}"})
    await notif.save_inapp(db, [user["id"]], "Shift Claimed ✓", f"You claimed {s['role']} on {s['start'][:10]}.", category="shift")
    return {"status": "claimed", "shift_id": new_shift["id"]}


@api.post("/open-shifts/{shift_id}/cancel-claim")
async def cancel_claim(shift_id: str, user=Depends(get_current_user)):
    await db.open_shifts.update_one({"id": shift_id}, {"$pull": {"claimed_by": user["id"], "waitlist": user["id"]}})
    await db.shifts.delete_one({"user_id": user["id"], "claimed_from_open": shift_id})
    # Auto-promote first person on waitlist
    s = await db.open_shifts.find_one({"id": shift_id})
    if s:
        waitlist = s.get("waitlist") or []
        claimed = s.get("claimed_by") or []
        if waitlist and len(claimed) < s.get("spots_available", 1):
            next_uid = waitlist[0]
            await db.open_shifts.update_one(
                {"id": shift_id},
                {"$pull": {"waitlist": next_uid}, "$addToSet": {"claimed_by": next_uid}},
            )
            promoted = {
                "id": str(uuid.uuid4()), "user_id": next_uid, "site_id": s["site_id"],
                "start": s["start"], "end": s["end"], "role": s["role"],
                "pay_rate": s["pay_rate"], "status": "scheduled",
                "claimed_from_open": shift_id, "instructions_acknowledged": False,
            }
            await db.shifts.insert_one(promoted)
            await send_push([next_uid], {"title": "Spot Available!", "message": f"You've been promoted from the waitlist for a {s['role']} shift on {s['start'][:10]}"})
            await notif.save_inapp(db, [next_uid], "Waitlist Promoted 🎉", f"You got a spot for {s['role']} on {s['start'][:10]}.", category="shift")
    return {"status": "cancelled"}


@api.get("/my-claims")
async def my_claims(user=Depends(get_current_user)):
    shifts = await db.shifts.find(
        {"user_id": user["id"], "claimed_from_open": {"$exists": True}},
        {"_id": 0},
    ).sort("start", -1).to_list(50)
    return {"claims": shifts}


# ============================================================
# TIME CLOCK
# ============================================================
def haversine_m(lat1, lon1, lat2, lon2):
    import math
    R = 6371000
    p1 = math.radians(lat1); p2 = math.radians(lat2)
    dp = math.radians(lat2 - lat1); dl = math.radians(lon2 - lon1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2 * R * math.asin(math.sqrt(a))


@api.get("/timeclock/status")
async def timeclock_status(user=Depends(get_current_user)):
    active = await db.timeclock.find_one({"user_id": user["id"], "clock_out": None}, {"_id": 0, "selfie_in": 0, "selfie_out": 0})
    return {"active": active}


@api.post("/timeclock/clock-in")
async def clock_in(body: ClockInIn, user=Depends(get_current_user)):
    existing = await db.timeclock.find_one({"user_id": user["id"], "clock_out": None})
    if existing:
        raise HTTPException(400, "Already clocked in. Clock out first.")

    site = None
    sh = None
    geofence_ok = True
    distance_m = None
    if body.site_id:
        site = await db.sites.find_one({"id": body.site_id}, {"_id": 0})
    elif body.shift_id:
        sh = await db.shifts.find_one({"id": body.shift_id})
        if sh:
            site = await db.sites.find_one({"id": sh["site_id"]}, {"_id": 0})

    if site:
        distance_m = haversine_m(body.latitude, body.longitude, site["latitude"], site["longitude"])
        geofence_ok = distance_m <= site.get("geofence_radius_m", 200)

    # Detect late clock-in (> 15 min after shift start)
    late_clock_in = False
    if sh and sh.get("start"):
        try:
            shift_start_dt = datetime.fromisoformat(sh["start"].replace("Z", "+00:00"))
            if shift_start_dt.tzinfo is None:
                shift_start_dt = shift_start_dt.replace(tzinfo=timezone.utc)
            mins_late = (now_utc() - shift_start_dt).total_seconds() / 60
            late_clock_in = mins_late > 15
        except Exception:
            pass

    entry = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "shift_id": body.shift_id,
        "site_id": site["id"] if site else None,
        "clock_in": iso(now_utc()),
        "clock_in_lat": body.latitude,
        "clock_in_lng": body.longitude,
        "selfie_in": body.selfie_base64,
        "geofence_ok": geofence_ok,
        "geofence_distance_m": distance_m,
        "late_clock_in": late_clock_in,
        "clock_out": None,
        "breaks": [],
        "hours_worked": None,
    }
    await db.timeclock.insert_one(entry)
    entry.pop("selfie_in", None)
    entry.pop("_id", None)
    return {"entry": entry, "geofence_ok": geofence_ok, "distance_m": distance_m, "late_clock_in": late_clock_in}


@api.post("/timeclock/clock-out")
async def clock_out(body: ClockOutIn, user=Depends(get_current_user)):
    active = await db.timeclock.find_one({"user_id": user["id"], "clock_out": None})
    if not active:
        raise HTTPException(400, "Not currently clocked in")
    clock_out_dt = now_utc()
    clock_in_dt = datetime.fromisoformat(active["clock_in"].replace("Z", "+00:00"))
    total_seconds = (clock_out_dt - clock_in_dt).total_seconds()
    # Subtract break minutes
    break_seconds = 0
    for br in active.get("breaks", []):
        if br.get("end"):
            b_start = datetime.fromisoformat(br["start"].replace("Z", "+00:00"))
            b_end = datetime.fromisoformat(br["end"].replace("Z", "+00:00"))
            break_seconds += (b_end - b_start).total_seconds()
    hours = round(max(0, total_seconds - break_seconds) / 3600, 2)
    await db.timeclock.update_one(
        {"id": active["id"]},
        {"$set": {
            "clock_out": iso(clock_out_dt),
            "clock_out_lat": body.latitude,
            "clock_out_lng": body.longitude,
            "selfie_out": body.selfie_base64,
            "hours_worked": hours,
        }},
    )
    # Auto-complete the associated shift
    if active.get("shift_id"):
        await db.shifts.update_one(
            {"id": active["shift_id"], "status": "scheduled"},
            {"$set": {"status": "completed", "hours_worked": hours}},
        )
    return {"hours_worked": hours}


@api.post("/timeclock/break")
async def break_action(body: BreakIn, user=Depends(get_current_user)):
    active = await db.timeclock.find_one({"user_id": user["id"], "clock_out": None})
    if not active:
        raise HTTPException(400, "Not currently clocked in")
    breaks = active.get("breaks", [])
    if body.action == "start":
        if breaks and not breaks[-1].get("end"):
            raise HTTPException(400, "Already on break")
        breaks.append({"start": iso(now_utc()), "end": None})
    elif body.action == "end":
        if not breaks or breaks[-1].get("end"):
            raise HTTPException(400, "Not on break")
        breaks[-1]["end"] = iso(now_utc())
    else:
        raise HTTPException(400, "Invalid action")
    await db.timeclock.update_one({"id": active["id"]}, {"$set": {"breaks": breaks}})
    return {"breaks": breaks}


@api.get("/timeclock/history")
async def timeclock_history(user=Depends(get_current_user), limit: int = 20):
    entries = await db.timeclock.find(
        {"user_id": user["id"]},
        {"_id": 0, "selfie_in": 0, "selfie_out": 0},
    ).sort("clock_in", -1).to_list(limit)
    return {"entries": entries}


# ============================================================
# WALLET
# ============================================================
@api.get("/wallet")
async def wallet(user=Depends(get_current_user)):
    docs = await db.wallet_documents.find({"user_id": user["id"]}, {"_id": 0}).to_list(50)
    qr_payload = f"SKYHAWK|{user['employee_number']}|{user['id']}"
    return {
        "documents": docs,
        "employee": {
            "id": user["id"],
            "full_name": user["full_name"],
            "employee_number": user.get("employee_number"),
            "photo_url": user.get("photo_url"),
            "role": user["role"],
            "licence_number": user.get("licence_number"),
            "employment_status": user.get("employment_status"),
        },
        "qr_payload": qr_payload,
    }


# ============================================================
# ANNOUNCEMENTS
# ============================================================
@api.get("/announcements")
async def announcements(user=Depends(get_current_user)):
    items = await db.announcements.find({}, {"_id": 0}).sort("posted_at", -1).to_list(50)
    for a in items:
        a["read"] = user["id"] in (a.get("read_by") or [])
        a["read_count"] = len(a.get("read_by") or [])
    return {"announcements": items}


@api.post("/announcements/{ann_id}/acknowledge")
async def ack_announcement(ann_id: str, user=Depends(get_current_user)):
    r = await db.announcements.update_one(
        {"id": ann_id}, {"$addToSet": {"read_by": user["id"]}}
    )
    if r.matched_count == 0:
        raise HTTPException(404, "Announcement not found")
    return {"acknowledged": True}


# ============================================================
# INCIDENTS
# ============================================================
@api.post("/incidents")
async def create_incident(body: IncidentIn, user=Depends(get_current_user)):
    inc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "user_name": user["full_name"],
        "type": body.type,
        "site_id": body.site_id,
        "description": body.description,
        "severity": body.severity,
        "witness_name": body.witness_name,
        "witness_contact": body.witness_contact,
        "photos": body.photos,
        "signature_base64": body.signature_base64,
        "status": "submitted",
        "created_at": iso(now_utc()),
    }
    await db.incidents.insert_one(inc)
    inc.pop("_id", None)
    inc.pop("photos", None)
    inc.pop("signature_base64", None)
    return inc


@api.get("/incidents")
async def list_incidents(user=Depends(get_current_user)):
    items = await db.incidents.find(
        {"user_id": user["id"]},
        {"_id": 0, "photos": 0, "signature_base64": 0},
    ).sort("created_at", -1).to_list(50)
    return {"incidents": items}


# ============================================================
# PAYROLL
# ============================================================
@api.get("/payroll")
async def payroll(user=Depends(get_current_user)):
    periods = await db.payroll.find({"user_id": user["id"]}, {"_id": 0}).sort("period_end", -1).to_list(20)
    current = periods[0] if periods else None
    total_hours = sum(p.get("hours_regular", 0) + p.get("hours_overtime", 0) for p in periods)
    total_gross = sum(p.get("gross", 0) for p in periods)
    return {"periods": periods, "current": current, "total_hours": total_hours, "total_gross": total_gross}


# ============================================================
# PROFILE
# ============================================================
@api.get("/profile")
async def profile(user=Depends(get_current_user)):
    equipment = await db.equipment.find({"user_id": user["id"]}, {"_id": 0}).to_list(20)
    return {"user": user, "equipment": equipment}


# ============================================================
# ONBOARDING
# ============================================================
ONBOARDING_STEPS = [
    "documents_uploaded",
    "sin_submitted",
    "direct_deposit_submitted",
    "emergency_contact_added",
    "agreements_signed",
]


async def _onboarding_status(user: dict) -> dict:
    ob = await db.onboarding.find_one({"user_id": user["id"]}, {"_id": 0})
    if not ob:
        ob = {"user_id": user["id"]}
    # Emergency contact can also be satisfied by the user profile field.
    if not ob.get("emergency_contact_added") and user.get("emergency_contact"):
        ob["emergency_contact_added"] = True
    status = {step: bool(ob.get(step)) for step in ONBOARDING_STEPS}
    completed = sum(1 for v in status.values() if v)
    total = len(ONBOARDING_STEPS)
    return {"status": status, "completed": completed, "total": total, "percent": int(completed / total * 100)}


@api.get("/onboarding/status")
async def onboarding_status(user=Depends(get_current_user)):
    return await _onboarding_status(user)


@api.post("/onboarding/documents", status_code=201)
async def onboarding_documents(body: OnboardingDocIn, user=Depends(get_current_user)):
    doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "type": body.type,
        "name": body.name,
        "number": body.number,
        "expiry": body.expiry,
        "status": _compliance_status(body.expiry),
        "attachment_url": body.attachment_url,
        "created_at": iso(now_utc()),
    }
    await db.wallet_documents.insert_one(doc)
    await db.onboarding.update_one(
        {"user_id": user["id"]},
        {"$set": {"documents_uploaded": True, "updated_at": iso(now_utc())}},
        upsert=True,
    )
    return await _onboarding_status(user)


@api.post("/onboarding/sin", status_code=201)
async def onboarding_sin(body: OnboardingSINIn, user=Depends(get_current_user)):
    await db.users.update_one({"id": user["id"]}, {"$set": {"sin": body.sin}})
    await db.onboarding.update_one(
        {"user_id": user["id"]},
        {"$set": {"sin_submitted": True, "updated_at": iso(now_utc())}},
        upsert=True,
    )
    return await _onboarding_status(user)


@api.post("/onboarding/direct-deposit", status_code=201)
async def onboarding_direct_deposit(body: OnboardingDirectDepositIn, user=Depends(get_current_user)):
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"direct_deposit": {"institution": body.institution, "transit": body.transit, "account": body.account}}},
    )
    await db.onboarding.update_one(
        {"user_id": user["id"]},
        {"$set": {"direct_deposit_submitted": True, "updated_at": iso(now_utc())}},
        upsert=True,
    )
    return await _onboarding_status(user)


@api.post("/onboarding/emergency-contact", status_code=201)
async def onboarding_emergency_contact(body: OnboardingEmergencyContactIn, user=Depends(get_current_user)):
    contact = {"name": body.name, "phone": body.phone, "relation": body.relation}
    await db.users.update_one({"id": user["id"]}, {"$set": {"emergency_contact": contact}})
    await db.onboarding.update_one(
        {"user_id": user["id"]},
        {"$set": {"emergency_contact_added": True, "updated_at": iso(now_utc())}},
        upsert=True,
    )
    return await _onboarding_status(user)


@api.post("/onboarding/agreements", status_code=201)
async def onboarding_agreements(body: OnboardingAgreementsIn, user=Depends(get_current_user)):
    if not body.signed:
        raise HTTPException(400, "Agreements must be signed")
    await db.onboarding.update_one(
        {"user_id": user["id"]},
        {"$set": {"agreements_signed": True, "updated_at": iso(now_utc())}},
        upsert=True,
    )
    # If all steps are complete, flip the master flag.
    status = await _onboarding_status(user)
    if status["percent"] == 100:
        await db.users.update_one({"id": user["id"]}, {"$set": {"onboarding_complete": True}})
    return status


# ============================================================
# PUSH NOTIFICATIONS
# ============================================================
@api.post("/register-push", status_code=201)
async def register_push(body: RegisterPushIn):
    try:
        resp = await push_client.post("/api/v1/push/users/register", json=body.model_dump())
        if resp.status_code == 401:
            # Placeholder key - store locally
            await db.push_tokens.update_one(
                {"user_id": body.user_id, "device_token": body.device_token},
                {"$set": body.model_dump()},
                upsert=True,
            )
            return {"status": "registered_local"}
        if resp.status_code >= 500:
            raise HTTPException(502, "Push provider unavailable")
        resp.raise_for_status()
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"Push register fallback: {e}")
        await db.push_tokens.update_one(
            {"user_id": body.user_id, "device_token": body.device_token},
            {"$set": body.model_dump()},
            upsert=True,
        )
    return {"status": "registered"}



# ============================================================
# ADMIN — Dependency & Models
# ============================================================

async def require_admin(user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Admin access required")
    return user


class AdminCreateGuardIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    full_name: str
    phone: Optional[str] = None
    employee_number: Optional[str] = None
    licence_number: Optional[str] = None
    licence_expiry: Optional[str] = None
    certifications: List[str] = []
    employment_status: str = "Active - Full Time"


class AdminUpdateGuardIn(BaseModel):
    full_name: Optional[str] = None
    phone: Optional[str] = None
    employee_number: Optional[str] = None
    licence_number: Optional[str] = None
    licence_expiry: Optional[str] = None
    certifications: Optional[List[str]] = None
    employment_status: Optional[str] = None
    onboarding_complete: Optional[bool] = None


class AdminCreateSiteIn(BaseModel):
    name: str
    address: str
    latitude: float
    longitude: float
    supervisor: str
    supervisor_phone: str
    instructions: str = ""
    geofence_radius_m: int = 150


class AdminUpdateSiteIn(BaseModel):
    name: Optional[str] = None
    address: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    supervisor: Optional[str] = None
    supervisor_phone: Optional[str] = None
    instructions: Optional[str] = None
    geofence_radius_m: Optional[int] = None


class AdminCreateShiftIn(BaseModel):
    user_id: str
    site_id: str
    start: str
    end: str
    role: str = "Patrol"
    pay_rate: float = 24.50
    notes: Optional[str] = None


class AdminUpdateShiftIn(BaseModel):
    user_id: Optional[str] = None
    site_id: Optional[str] = None
    start: Optional[str] = None
    end: Optional[str] = None
    role: Optional[str] = None
    pay_rate: Optional[float] = None
    status: Optional[str] = None
    hours_worked: Optional[float] = None
    notes: Optional[str] = None


class AdminCreateOpenShiftIn(BaseModel):
    site_id: str
    start: str
    end: str
    role: str = "Patrol"
    pay_rate: float = 24.50
    spots_available: int = 1
    urgent: bool = False


class AdminCreateAnnouncementIn(BaseModel):
    title: str
    body: str
    severity: str = "info"
    site_id: Optional[str] = None
    posted_by: Optional[str] = None


class AdminUpdateIncidentIn(BaseModel):
    status: str
    notes: Optional[str] = None


class DeductionItem(BaseModel):
    label: str
    amount: float


class AdminCreatePayrollIn(BaseModel):
    user_id: str
    period_start: str
    period_end: str
    pay_date: Optional[str] = None          # defaults to period_end if omitted
    hours_regular: Optional[float] = None   # kept for back-compat
    hours_worked: Optional[float] = None    # frontend-friendly alias
    hours_overtime: float = 0.0
    pay_rate: Optional[float] = None        # kept for back-compat
    hourly_rate: Optional[float] = None     # frontend-friendly alias
    status: str = "submitted"
    notes: Optional[str] = None
    deductions: List[DeductionItem] = []
    paid_via: Optional[str] = None          # cash | cheque | direct_deposit | interac


class AdminUpdatePayrollIn(BaseModel):
    status: Optional[str] = None
    hours_regular: Optional[float] = None
    hours_worked: Optional[float] = None
    hours_overtime: Optional[float] = None
    pay_rate: Optional[float] = None
    hourly_rate: Optional[float] = None
    pay_date: Optional[str] = None
    gross: Optional[float] = None
    net: Optional[float] = None
    notes: Optional[str] = None
    deductions: Optional[List[DeductionItem]] = None
    paid_via: Optional[str] = None


class AdminUpdateTimeclockIn(BaseModel):
    clock_in: Optional[str] = None
    clock_out: Optional[str] = None
    hours_worked: Optional[float] = None


# ============================================================
# ADMIN ROUTES — Dashboard
# ============================================================

@api.get("/admin/dashboard")
async def admin_dashboard(admin=Depends(require_admin)):
    now = now_utc()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)
    total_guards = await db.users.count_documents({"role": "employee"})
    shifts_today = await db.shifts.count_documents({"start": {"$gte": iso(today_start), "$lt": iso(today_end)}})
    active_clocked = await db.timeclock.count_documents({"clock_out": None})
    open_incidents = await db.incidents.count_documents({"status": "submitted"})
    pending_payroll = await db.payroll.count_documents({"status": {"$in": ["submitted", "under_review"]}})
    recent_incidents = await db.incidents.find(
        {}, {"_id": 0, "photos": 0, "signature_base64": 0}
    ).sort("created_at", -1).to_list(5)
    active_entries = await db.timeclock.find({"clock_out": None}, {"_id": 0, "selfie_in": 0}).to_list(20)
    if active_entries:
        uids = [e["user_id"] for e in active_entries]
        sids = [e["site_id"] for e in active_entries if e.get("site_id")]
        um = {u["id"]: u for u in await db.users.find({"id": {"$in": uids}}, {"_id": 0, "hashed_password": 0}).to_list(20)}
        sm = {s["id"]: s for s in await db.sites.find({"id": {"$in": sids}}, {"_id": 0}).to_list(20)} if sids else {}
        for e in active_entries:
            e["user"] = um.get(e["user_id"])
            e["site"] = sm.get(e.get("site_id"))
    return {
        "total_guards": total_guards, "shifts_today": shifts_today,
        "active_clocked": active_clocked, "open_incidents": open_incidents,
        "pending_payroll": pending_payroll, "recent_incidents": recent_incidents,
        "active_entries": active_entries,
    }


# ============================================================
# ADMIN ROUTES — Guards
# ============================================================

@api.get("/admin/guards")
async def admin_list_guards(admin=Depends(require_admin)):
    guards = await db.users.find({}, {"_id": 0, "hashed_password": 0}).sort("full_name", 1).to_list(500)
    return {"guards": guards}


@api.post("/admin/guards", status_code=201)
async def admin_create_guard(body: AdminCreateGuardIn, admin=Depends(require_admin)):
    if await db.users.find_one({"email": body.email.lower()}):
        raise HTTPException(400, "Email already registered")
    user_id = str(uuid.uuid4())
    guard = {
        "id": user_id, "email": body.email.lower(),
        "hashed_password": hash_password(body.password),
        "full_name": body.full_name, "phone": body.phone, "role": "employee",
        "employee_number": body.employee_number or f"SH-{str(uuid.uuid4())[:4].upper()}",
        "licence_number": body.licence_number, "licence_expiry": body.licence_expiry,
        "certifications": body.certifications, "employment_status": body.employment_status,
        "photo_url": None, "emergency_contact": None, "onboarding_complete": False,
        "created_at": iso(now_utc()),
    }
    await db.users.insert_one(guard)
    guard.pop("hashed_password", None); guard.pop("_id", None)
    return guard


@api.get("/admin/guards/{guard_id}")
async def admin_get_guard(guard_id: str, admin=Depends(require_admin)):
    guard = await db.users.find_one({"id": guard_id}, {"_id": 0, "hashed_password": 0})
    if not guard:
        raise HTTPException(404, "Guard not found")
    recent_shifts = await db.shifts.find({"user_id": guard_id}, {"_id": 0}).sort("start", -1).to_list(10)
    for s in recent_shifts:
        s["site"] = await db.sites.find_one({"id": s["site_id"]}, {"_id": 0})
    recent_clock = await db.timeclock.find(
        {"user_id": guard_id}, {"_id": 0, "selfie_in": 0, "selfie_out": 0}
    ).sort("clock_in", -1).to_list(10)
    equipment = await db.equipment.find({"user_id": guard_id}, {"_id": 0}).to_list(20)
    documents = await db.wallet_documents.find({"user_id": guard_id}, {"_id": 0}).sort("uploaded_at", -1).to_list(50)
    onboarding = await db.onboarding.find_one({"user_id": guard_id}, {"_id": 0})
    return {
        "guard": guard,
        "recent_shifts": recent_shifts,
        "recent_clock": recent_clock,
        "equipment": equipment,
        "documents": documents,
        "onboarding": onboarding,
    }


@api.put("/admin/guards/{guard_id}")
async def admin_update_guard(guard_id: str, body: AdminUpdateGuardIn, admin=Depends(require_admin)):
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(400, "No fields to update")
    r = await db.users.update_one({"id": guard_id}, {"$set": update})
    if r.matched_count == 0:
        raise HTTPException(404, "Guard not found")
    return await db.users.find_one({"id": guard_id}, {"_id": 0, "hashed_password": 0})


# ============================================================
# ADMIN ROUTES — Sites
# ============================================================

@api.get("/admin/sites")
async def admin_list_sites(admin=Depends(require_admin)):
    sites = await db.sites.find({}, {"_id": 0}).sort("name", 1).to_list(200)
    return {"sites": sites}


@api.post("/admin/sites", status_code=201)
async def admin_create_site(body: AdminCreateSiteIn, admin=Depends(require_admin)):
    site = {"id": str(uuid.uuid4()), **body.model_dump()}
    await db.sites.insert_one(site)
    site.pop("_id", None)
    return site


@api.put("/admin/sites/{site_id}")
async def admin_update_site(site_id: str, body: AdminUpdateSiteIn, admin=Depends(require_admin)):
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(400, "No fields to update")
    r = await db.sites.update_one({"id": site_id}, {"$set": update})
    if r.matched_count == 0:
        raise HTTPException(404, "Site not found")
    return await db.sites.find_one({"id": site_id}, {"_id": 0})


@api.delete("/admin/sites/{site_id}", status_code=204)
async def admin_delete_site(site_id: str, admin=Depends(require_admin)):
    r = await db.sites.delete_one({"id": site_id})
    if r.deleted_count == 0:
        raise HTTPException(404, "Site not found")


# ============================================================
# ADMIN ROUTES — Shifts
# ============================================================

@api.get("/admin/shifts")
async def admin_list_shifts(
    admin=Depends(require_admin),
    user_id: Optional[str] = None,
    site_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    shift_status: Optional[str] = Query(None, alias="status"),
):
    q: dict = {}
    if user_id: q["user_id"] = user_id
    if site_id: q["site_id"] = site_id
    if shift_status: q["status"] = shift_status
    if from_date or to_date:
        q["start"] = {}
        if from_date: q["start"]["$gte"] = from_date
        if to_date: q["start"]["$lte"] = to_date
    shifts = await db.shifts.find(q, {"_id": 0}).sort("start", -1).to_list(500)
    if shifts:
        sids = list({s["site_id"] for s in shifts})
        uids = list({s["user_id"] for s in shifts})
        sm = {s["id"]: s for s in await db.sites.find({"id": {"$in": sids}}, {"_id": 0}).to_list(200)}
        um = {u["id"]: u for u in await db.users.find({"id": {"$in": uids}}, {"_id": 0, "hashed_password": 0}).to_list(500)}
        for s in shifts:
            s["site"] = sm.get(s["site_id"]); s["user"] = um.get(s["user_id"])
    return {"shifts": shifts}


@api.post("/admin/shifts", status_code=201)
async def admin_create_shift(body: AdminCreateShiftIn, admin=Depends(require_admin)):
    if not await db.users.find_one({"id": body.user_id}):
        raise HTTPException(404, "Guard not found")
    if not await db.sites.find_one({"id": body.site_id}):
        raise HTTPException(404, "Site not found")
    # Conflict detection: guard must not have an overlapping scheduled shift
    overlap = await db.shifts.find_one({
        "user_id": body.user_id,
        "status": {"$ne": "cancelled"},
        "start": {"$lt": body.end},
        "end": {"$gt": body.start},
    })
    if overlap:
        raise HTTPException(409, f"Scheduling conflict: guard already has a shift from {overlap['start'][:16]} to {overlap['end'][:16]}.")
    shift = {
        "id": str(uuid.uuid4()), "user_id": body.user_id, "site_id": body.site_id,
        "start": body.start, "end": body.end, "role": body.role, "pay_rate": body.pay_rate,
        "notes": body.notes,
        "status": "scheduled", "instructions_acknowledged": False,
        "created_by": admin["id"], "created_at": iso(now_utc()),
    }
    await db.shifts.insert_one(shift); shift.pop("_id", None)
    await send_push([body.user_id], {"title": "New Shift Assigned", "message": f"{body.role} shift on {body.start[:10]}"})
    return shift


@api.put("/admin/shifts/{shift_id}")
async def admin_update_shift(shift_id: str, body: AdminUpdateShiftIn, admin=Depends(require_admin)):
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(400, "No fields to update")
    r = await db.shifts.update_one({"id": shift_id}, {"$set": update})
    if r.matched_count == 0:
        raise HTTPException(404, "Shift not found")
    return await db.shifts.find_one({"id": shift_id}, {"_id": 0})


@api.delete("/admin/shifts/{shift_id}", status_code=204)
async def admin_delete_shift(shift_id: str, admin=Depends(require_admin)):
    r = await db.shifts.delete_one({"id": shift_id})
    if r.deleted_count == 0:
        raise HTTPException(404, "Shift not found")


# ============================================================
# ADMIN ROUTES — Open Shifts
# ============================================================

@api.get("/admin/open-shifts")
async def admin_list_open_shifts(admin=Depends(require_admin)):
    shifts = await db.open_shifts.find({}, {"_id": 0}).sort("start", 1).to_list(200)
    if shifts:
        sids = list({s["site_id"] for s in shifts})
        sm = {s["id"]: s for s in await db.sites.find({"id": {"$in": sids}}, {"_id": 0}).to_list(50)}
        for s in shifts:
            s["site"] = sm.get(s["site_id"]); s["claimed_count"] = len(s.get("claimed_by") or [])
    return {"shifts": shifts}


@api.post("/admin/open-shifts", status_code=201)
async def admin_create_open_shift(body: AdminCreateOpenShiftIn, admin=Depends(require_admin)):
    if not await db.sites.find_one({"id": body.site_id}):
        raise HTTPException(404, "Site not found")
    shift = {"id": str(uuid.uuid4()), **body.model_dump(), "posted_at": iso(now_utc()), "claimed_by": [], "waitlist": [], "posted_by": admin["id"]}
    await db.open_shifts.insert_one(shift); shift.pop("_id", None)
    return shift


@api.delete("/admin/open-shifts/{shift_id}", status_code=204)
async def admin_delete_open_shift(shift_id: str, admin=Depends(require_admin)):
    r = await db.open_shifts.delete_one({"id": shift_id})
    if r.deleted_count == 0:
        raise HTTPException(404, "Open shift not found")


# ============================================================
# ADMIN ROUTES — Announcements
# ============================================================

@api.get("/admin/announcements")
async def admin_list_announcements(admin=Depends(require_admin)):
    items = await db.announcements.find({}, {"_id": 0}).sort("posted_at", -1).to_list(200)
    total_guards = await db.users.count_documents({"role": "employee"})
    for a in items:
        a["read_count"] = len(a.get("read_by") or []); a["total_guards"] = total_guards
    return {"announcements": items}


@api.post("/admin/announcements", status_code=201)
async def admin_create_announcement(body: AdminCreateAnnouncementIn, admin=Depends(require_admin)):
    ann = {
        "id": str(uuid.uuid4()), "title": body.title, "body": body.body,
        "severity": body.severity, "site_id": body.site_id,
        "posted_by": body.posted_by or admin.get("full_name", "Admin"),
        "posted_at": iso(now_utc()), "read_by": [],
    }
    await db.announcements.insert_one(ann); ann.pop("_id", None)
    guards = await db.users.find({"role": "employee"}, {"id": 1, "email": 1, "_id": 0}).to_list(500)
    guard_ids = [g["id"] for g in guards]
    await send_push(guard_ids, {"title": body.title, "message": body.body[:120]})
    ann_html = notif.email_template(
        body.title,
        [body.body],
        cta_text="View in App",
        cta_link="skyhawksecurity://announcements",
    )
    await asyncio.gather(
        *[notif.send_email(g["email"], f"[Skyhawk] {body.title}", ann_html, body.body) for g in guards if g.get("email")],
        notif.save_inapp(db, guard_ids, body.title, body.body[:300], category="announcement"),
        return_exceptions=True,
    )
    return ann


@api.delete("/admin/announcements/{ann_id}", status_code=204)
async def admin_delete_announcement(ann_id: str, admin=Depends(require_admin)):
    r = await db.announcements.delete_one({"id": ann_id})
    if r.deleted_count == 0:
        raise HTTPException(404, "Announcement not found")


# ============================================================
# ADMIN ROUTES — Incidents
# ============================================================

@api.get("/admin/incidents")
async def admin_list_incidents(
    admin=Depends(require_admin),
    inc_status: Optional[str] = Query(None, alias="status"),
):
    q: dict = {}
    if inc_status: q["status"] = inc_status
    items = await db.incidents.find(q, {"_id": 0, "photos": 0, "signature_base64": 0}).sort("created_at", -1).to_list(200)
    if items:
        sids = list({i["site_id"] for i in items if i.get("site_id")})
        if sids:
            sm = {s["id"]: s for s in await db.sites.find({"id": {"$in": sids}}, {"_id": 0}).to_list(50)}
            for inc in items:
                inc["site"] = sm.get(inc.get("site_id"))
    return {"incidents": items}


@api.put("/admin/incidents/{inc_id}")
async def admin_update_incident(inc_id: str, body: AdminUpdateIncidentIn, admin=Depends(require_admin)):
    update: dict = {"status": body.status, "reviewed_by": admin["id"], "reviewed_at": iso(now_utc())}
    if body.notes: update["review_notes"] = body.notes
    r = await db.incidents.update_one({"id": inc_id}, {"$set": update})
    if r.matched_count == 0:
        raise HTTPException(404, "Incident not found")
    return await db.incidents.find_one({"id": inc_id}, {"_id": 0, "photos": 0, "signature_base64": 0})


# ============================================================
# ADMIN ROUTES — Payroll
# ============================================================

@api.get("/admin/payroll")
async def admin_list_payroll(
    admin=Depends(require_admin),
    user_id: Optional[str] = None,
    payroll_status: Optional[str] = Query(None, alias="status"),
):
    q: dict = {}
    if user_id: q["user_id"] = user_id
    if payroll_status: q["status"] = payroll_status
    periods = await db.payroll.find(q, {"_id": 0}).sort("period_end", -1).to_list(500)
    if periods:
        uids = list({p["user_id"] for p in periods})
        um = {u["id"]: u for u in await db.users.find({"id": {"$in": uids}}, {"_id": 0, "hashed_password": 0}).to_list(500)}
        for p in periods: p["user"] = um.get(p["user_id"])
    return {"periods": periods}


@api.post("/admin/payroll", status_code=201)
async def admin_create_payroll(body: AdminCreatePayrollIn, admin=Depends(require_admin)):
    hours_reg = body.hours_regular if body.hours_regular is not None else (body.hours_worked or 0.0)
    rate = body.pay_rate if body.pay_rate is not None else (body.hourly_rate or 0.0)
    gross = round(hours_reg * rate + body.hours_overtime * rate * 1.5, 2)
    total_deductions = round(sum(d.amount for d in body.deductions), 2)
    net = round(gross - total_deductions, 2)
    record = {
        "id": str(uuid.uuid4()), "user_id": body.user_id,
        "period_start": body.period_start, "period_end": body.period_end,
        "pay_date": body.pay_date or body.period_end,
        "hours_regular": hours_reg, "hours_overtime": body.hours_overtime,
        "pay_rate": rate, "gross": gross,
        "deductions": [d.model_dump() for d in body.deductions],
        "total_deductions": total_deductions,
        "net": net,
        "status": body.status, "notes": body.notes or "",
        "paid_via": body.paid_via,
        "pay_stub_url": None,
        "created_by": admin["id"], "created_at": iso(now_utc()),
    }
    await db.payroll.insert_one(record); record.pop("_id", None)
    return record


# ── Timesheet Import ────────────────────────────────────────────────────────

def _parse_timesheet_xlsx(data: bytes, from_date=None, to_date=None):
    """Parse a client timesheet xlsx and return per-guard shift summaries."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not available on server. Contact admin.")

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    guards: dict = {}
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 10)]
        if not any(row):
            continue
        name_raw, lic_raw, date_val, project = row[0], row[1], row[2], row[3]
        hours = float(row[7]) if row[7] is not None else 0.0

        if not name_raw:
            continue
        name_str = str(name_raw).strip()
        if name_str.lower().startswith("skyhawk"):
            continue

        # Parse shift date
        shift_date = None
        if isinstance(date_val, datetime):
            shift_date = date_val.date()
        elif isinstance(date_val, date):
            shift_date = date_val

        # Optional date-range filter
        if shift_date and from_date and shift_date < from_date:
            continue
        if shift_date and to_date and shift_date > to_date:
            continue

        # Parse licence number (stored as float in some cells)
        if lic_raw is not None:
            try:
                lic_str = str(int(float(lic_raw)))
            except Exception:
                lic_str = str(lic_raw).strip()
        else:
            m = re.search(r"(\d{7,10})$", name_str)
            lic_str = m.group(1) if m else None
        if not lic_str:
            continue

        # Clean name – strip {N} repeat-markers and embedded lic number at tail
        clean = re.sub(r"\s*\{[^}]*\}", "", name_str).strip()
        clean = re.sub(r"\s+\d{6,10}$", "", clean).strip()

        if lic_str not in guards:
            guards[lic_str] = {
                "lic_number": lic_str, "guard_name": clean,
                "total_hours": 0.0, "shift_count": 0,
                "dates": [], "projects": [],
            }
        g = guards[lic_str]
        g["total_hours"] += hours
        g["shift_count"] += 1
        if shift_date:
            g["dates"].append(shift_date.isoformat())
        proj = str(project).strip() if project else None
        if proj and proj not in g["projects"]:
            g["projects"].append(proj)

    result = []
    for g in guards.values():
        ds = sorted(g["dates"])
        result.append({
            "lic_number":   g["lic_number"],
            "guard_name":   g["guard_name"],
            "total_hours":  round(g["total_hours"], 2),
            "shift_count":  g["shift_count"],
            "period_start": ds[0]  if ds else None,
            "period_end":   ds[-1] if ds else None,
            "projects":     g["projects"][:5],
        })
    result.sort(key=lambda x: x["guard_name"].upper())
    return result


@api.post("/admin/payroll/parse-timesheet")
async def parse_timesheet_upload(
    file: UploadFile = File(...),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    admin=Depends(require_admin),
):
    data = await file.read()
    from_d = date.fromisoformat(from_date) if from_date else None
    to_d   = date.fromisoformat(to_date)   if to_date   else None
    try:
        guards = _parse_timesheet_xlsx(data, from_d, to_d)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, f"Could not parse file: {exc}")

    if not guards:
        raise HTTPException(400, "No valid shift rows found. Check the date range or file format.")

    # Match guards in DB by licence_number field
    lic_numbers = [g["lic_number"] for g in guards]
    db_users = await db.users.find(
        {"licence_number": {"$in": lic_numbers}},
        {"_id": 0, "id": 1, "full_name": 1, "licence_number": 1},
    ).to_list(5000)
    lic_to_user = {u["licence_number"]: u for u in db_users}

    matched = 0
    for g in guards:
        u = lic_to_user.get(g["lic_number"])
        g["user_id"]      = u["id"]        if u else None
        g["matched_name"] = u["full_name"] if u else None
        g["matched"]      = bool(u)
        if u:
            matched += 1

    return {
        "guards":         guards,
        "total_guards":   len(guards),
        "total_shifts":   sum(g["shift_count"] for g in guards),
        "total_hours":    round(sum(g["total_hours"] for g in guards), 2),
        "matched_guards": matched,
    }


class BulkPayrollEntryIn(BaseModel):
    lic_number:   str
    guard_name:   str
    user_id:      Optional[str] = None
    period_start: str
    period_end:   str
    pay_date:     Optional[str] = None
    hours_regular: float
    pay_rate:     float
    notes:        Optional[str] = ""


@api.post("/admin/payroll/bulk-create", status_code=201)
async def bulk_create_payroll(entries: List[BulkPayrollEntryIn], admin=Depends(require_admin)):
    records = []
    for e in entries:
        gross = round(e.hours_regular * e.pay_rate, 2)
        records.append({
            "id":               str(uuid.uuid4()),
            "user_id":          e.user_id or f"ext:{e.lic_number}",
            "lic_number":       e.lic_number,
            "guard_name_import": e.guard_name,
            "period_start":     e.period_start,
            "period_end":       e.period_end,
            "pay_date":         e.pay_date or e.period_end,
            "hours_regular":    e.hours_regular,
            "hours_overtime":   0.0,
            "pay_rate":         e.pay_rate,
            "gross":            gross,
            "deductions":       [],
            "total_deductions": 0.0,
            "net":              gross,
            "status":           "submitted",
            "notes":            e.notes or "",
            "paid_via":         None,
            "pay_stub_url":     None,
            "imported":         True,
            "created_by":       admin["id"],
            "created_at":       iso(now_utc()),
        })
    if records:
        await db.payroll.insert_many(records)
        for r in records:
            r.pop("_id", None)
    return {"created": len(records)}


@api.put("/admin/payroll/{payroll_id}")
async def admin_update_payroll(payroll_id: str, body: AdminUpdatePayrollIn, admin=Depends(require_admin)):
    raw = body.model_dump()
    # Normalise aliased fields
    if raw.get("hours_worked") is not None and raw.get("hours_regular") is None:
        raw["hours_regular"] = raw["hours_worked"]
    if raw.get("hourly_rate") is not None and raw.get("pay_rate") is None:
        raw["pay_rate"] = raw["hourly_rate"]
    raw.pop("hours_worked", None); raw.pop("hourly_rate", None)
    # Serialise deduction items
    if raw.get("deductions") is not None:
        raw["deductions"] = [d if isinstance(d, dict) else d.model_dump() for d in raw["deductions"]]
        raw["total_deductions"] = round(sum(d["amount"] for d in raw["deductions"]), 2)
    update = {k: v for k, v in raw.items() if v is not None}
    if not update:
        raise HTTPException(400, "No fields to update")
    r = await db.payroll.update_one({"id": payroll_id}, {"$set": update})
    if r.matched_count == 0:
        raise HTTPException(404, "Payroll record not found")
    updated = await db.payroll.find_one({"id": payroll_id}, {"_id": 0})
    if updated and updated.get("status") == "released":
        await send_push([updated["user_id"]], {"title": "Payroll Released", "message": f"Pay for {updated.get('period_end','')[:10]} has been released."})
        guard = await db.users.find_one({"id": updated["user_id"]}, {"_id": 0, "email": 1, "full_name": 1})
        pay_html = notif.email_template(
            "Your pay has been released",
            [
                f"Hi {guard.get('full_name', 'there') if guard else 'there'},",
                f"Your pay for the period ending <strong>{updated.get('period_end','')[:10]}</strong> has been released.",
                f"Net pay: <strong>${updated.get('net', 0):.2f}</strong> (Gross: ${updated.get('gross', 0):.2f})",
            ],
            cta_text="View Pay Stub",
            cta_link="skyhawksecurity://wallet",
        )
        await asyncio.gather(
            notif.send_email(guard["email"], "Your Skyhawk pay has been released", pay_html) if guard and guard.get("email") else asyncio.sleep(0),
            notif.save_inapp(db, [updated["user_id"]], "Payroll Released 💰", f"Pay for {updated.get('period_end','')[:10]} released. Net: ${updated.get('net', 0):.2f}", category="payroll"),
            return_exceptions=True,
        )
    return updated


# ============================================================
# ADMIN ROUTES — Timeclock
# ============================================================

@api.get("/admin/timeclock")
async def admin_list_timeclock(
    admin=Depends(require_admin),
    user_id: Optional[str] = None,
    active_only: bool = False,
):
    q: dict = {}
    if user_id: q["user_id"] = user_id
    if active_only: q["clock_out"] = None
    entries = await db.timeclock.find(q, {"_id": 0, "selfie_in": 0, "selfie_out": 0}).sort("clock_in", -1).to_list(200)
    if entries:
        uids = list({e["user_id"] for e in entries})
        sids = list({e["site_id"] for e in entries if e.get("site_id")})
        um = {u["id"]: u for u in await db.users.find({"id": {"$in": uids}}, {"_id": 0, "hashed_password": 0}).to_list(200)}
        sm = {s["id"]: s for s in await db.sites.find({"id": {"$in": sids}}, {"_id": 0}).to_list(50)} if sids else {}
        for e in entries:
            e["user"] = um.get(e["user_id"]); e["site"] = sm.get(e.get("site_id"))
    return {"entries": entries}


@api.put("/admin/timeclock/{entry_id}")
async def admin_update_timeclock(entry_id: str, body: AdminUpdateTimeclockIn, admin=Depends(require_admin)):
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(400, "No fields to update")
    update.update({"manually_adjusted": True, "adjusted_by": admin["id"], "adjusted_at": iso(now_utc())})
    r = await db.timeclock.update_one({"id": entry_id}, {"$set": update})
    if r.matched_count == 0:
        raise HTTPException(404, "Timeclock entry not found")
    return await db.timeclock.find_one({"id": entry_id}, {"_id": 0, "selfie_in": 0, "selfie_out": 0})


# ============================================================
# SOS / PANIC
# ============================================================

@api.post("/sos", status_code=201)
async def trigger_sos(body: SOSIn, user=Depends(get_current_user)):
    alert_id = str(uuid.uuid4())
    now = now_utc()
    alert = {
        "id": alert_id,
        "user_id": user["id"],
        "user_name": user["full_name"],
        "employee_number": user.get("employee_number"),
        "latitude": body.latitude,
        "longitude": body.longitude,
        "message": body.message,
        "status": "active",
        "created_at": iso(now),
        "acknowledged_by": None,
        "acknowledged_at": None,
    }
    await db.sos_alerts.insert_one(alert)
    alert.pop("_id", None)
    incident = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "user_name": user["full_name"],
        "type": "incident",
        "site_id": None,
        "description": f"SOS ALERT triggered by {user['full_name']}. GPS: {body.latitude:.5f},{body.longitude:.5f}. {body.message or ''}".strip(),
        "severity": "critical",
        "witness_name": None,
        "witness_contact": None,
        "photos": [],
        "signature_base64": None,
        "status": "open",
        "sos_alert_id": alert_id,
        "created_at": iso(now),
        "audit_trail": [],
    }
    await db.incidents.insert_one(incident)
    admins = await db.users.find({"role": "admin"}, {"id": 1, "email": 1, "phone": 1, "_id": 0}).to_list(100)
    admin_ids = [a["id"] for a in admins]
    await send_push(
        admin_ids,
        {"title": f"🆘 SOS — {user['full_name']}", "message": f"{body.message or 'Guard needs immediate assistance'} · {body.latitude:.4f},{body.longitude:.4f}"},
        idempotency_key=f"sos-{alert_id}",
    )
    # Email + SMS all admins — SOS is critical, use every channel
    sos_html = notif.email_template(
        f"🆘 SOS ALERT — {user['full_name']}",
        [
            f"<strong style='color:#EF4444'>EMERGENCY — Guard needs immediate assistance.</strong>",
            f"<strong>Guard:</strong> {user['full_name']} (#{user.get('employee_number', 'N/A')})",
            f"<strong>GPS:</strong> {body.latitude:.5f}, {body.longitude:.5f}",
            f"<strong>Message:</strong> {body.message or 'No message provided'}",
            f"<strong>Time:</strong> {iso(now)}",
        ],
    )
    sms_body = (
        f"🆘 SOS ALERT: {user['full_name']} needs help. "
        f"GPS: {body.latitude:.4f},{body.longitude:.4f}. "
        f"{body.message or 'Immediate assistance required.'}"
    )
    await asyncio.gather(
        *[notif.send_email(a["email"], f"🆘 SOS ALERT — {user['full_name']}", sos_html) for a in admins if a.get("email")],
        *[notif.send_sms(a["phone"], sms_body) for a in admins if a.get("phone")],
        notif.save_inapp(db, admin_ids, f"🆘 SOS — {user['full_name']}", body.message or "Guard needs immediate assistance", category="sos"),
        notif.save_inapp(db, [user["id"]], "SOS Alert Sent", "Your emergency alert has been sent to all managers.", category="sos"),
        return_exceptions=True,
    )
    logger.warning("SOS triggered by user %s at %s,%s", user["id"], body.latitude, body.longitude)
    return alert


@api.get("/sos/active")
async def list_sos_active(admin=Depends(require_admin)):
    alerts = await db.sos_alerts.find(
        {"status": {"$in": ["active", "acknowledged"]}}, {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    uids = list({a["user_id"] for a in alerts})
    if uids:
        um = {u["id"]: u for u in await db.users.find({"id": {"$in": uids}}, {"_id": 0, "hashed_password": 0}).to_list(50)}
        for a in alerts:
            a["user"] = um.get(a["user_id"])
    return {"alerts": alerts}


@api.get("/sos/history")
async def sos_history(admin=Depends(require_admin)):
    alerts = await db.sos_alerts.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return {"alerts": alerts}


@api.post("/sos/{alert_id}/acknowledge")
async def acknowledge_sos(alert_id: str, admin=Depends(require_admin)):
    r = await db.sos_alerts.update_one(
        {"id": alert_id, "status": "active"},
        {"$set": {"status": "acknowledged", "acknowledged_by": admin["id"], "acknowledged_at": iso(now_utc())}},
    )
    if r.matched_count == 0:
        raise HTTPException(404, "Alert not found or already acknowledged")
    alert = await db.sos_alerts.find_one({"id": alert_id}, {"_id": 0})
    if alert:
        await send_push(
            [alert["user_id"]],
            {"title": "SOS Acknowledged", "message": f"Help is on the way. Acknowledged by {admin['full_name']}."},
        )
    return {"acknowledged": True}


@api.post("/sos/{alert_id}/resolve")
async def resolve_sos(alert_id: str, admin=Depends(require_admin)):
    r = await db.sos_alerts.update_one(
        {"id": alert_id},
        {"$set": {"status": "resolved", "resolved_at": iso(now_utc()), "resolved_by": admin["id"]}},
    )
    if r.matched_count == 0:
        raise HTTPException(404, "Alert not found")
    return {"resolved": True}


# ============================================================
# INCIDENT LIFECYCLE
# ============================================================

_INCIDENT_STATUSES = {"submitted", "open", "under_review", "escalated", "resolved"}


@api.patch("/incidents/{inc_id}/status")
async def update_incident_status(inc_id: str, body: IncidentStatusIn, admin=Depends(require_admin)):
    if body.status not in _INCIDENT_STATUSES:
        raise HTTPException(400, f"Invalid status. Must be one of: {', '.join(_INCIDENT_STATUSES)}")
    inc = await db.incidents.find_one({"id": inc_id}, {"_id": 0, "photos": 0, "signature_base64": 0})
    if not inc:
        raise HTTPException(404, "Incident not found")
    now = now_utc()
    update: dict = {"status": body.status, "updated_at": iso(now), "last_updated_by": admin["id"]}
    if body.assigned_to:
        update["assigned_to"] = body.assigned_to
    note_entry = {"ts": iso(now), "by": admin["full_name"], "status": body.status, "note": body.note}
    await db.incidents.update_one({"id": inc_id}, {"$set": update, "$push": {"audit_trail": note_entry}})
    if body.status == "resolved":
        msg = f"Your {inc['type'].replace('_', ' ')} report has been resolved. {body.note or ''}".strip()
        await send_push([inc["user_id"]], {"title": "Incident Resolved", "message": msg})
        await notif.save_inapp(db, [inc["user_id"]], "Incident Resolved ✓", msg, category="incident")
    elif body.status in ("escalated", "under_review"):
        label = "Escalated" if body.status == "escalated" else "Under Review"
        await notif.save_inapp(db, [inc["user_id"]], f"Incident {label}", f"Your report is now {body.status.replace('_',' ')}. {body.note or ''}".strip(), category="incident")
    return await db.incidents.find_one({"id": inc_id}, {"_id": 0, "photos": 0, "signature_base64": 0})


@api.get("/incidents/{inc_id}")
async def get_incident(inc_id: str, user=Depends(get_current_user)):
    q: dict = {"id": inc_id}
    if user["role"] != "admin":
        q["user_id"] = user["id"]
    inc = await db.incidents.find_one(q, {"_id": 0, "photos": 0, "signature_base64": 0})
    if not inc:
        raise HTTPException(404, "Incident not found")
    return inc


# ============================================================
# LIVE GPS TRACKING
# ============================================================

@api.post("/timeclock/location-ping")
async def location_ping(body: LocationPingIn, user=Depends(get_current_user)):
    active = await db.timeclock.find_one({"user_id": user["id"], "clock_out": None})
    if not active:
        raise HTTPException(400, "Not currently clocked in")
    ping = {
        "timeclock_id": active["id"],
        "user_id": user["id"],
        "latitude": body.latitude,
        "longitude": body.longitude,
        "accuracy_m": body.accuracy_m,
        "ts": iso(now_utc()),
    }
    await db.location_pings.insert_one(ping)
    await db.timeclock.update_one(
        {"id": active["id"]},
        {"$set": {"last_ping_lat": body.latitude, "last_ping_lng": body.longitude, "last_ping_ts": iso(now_utc())}},
    )
    return {"ok": True}


@api.get("/ops/live-locations")
async def live_locations(admin=Depends(require_admin)):
    """All guards currently clocked in with their last GPS ping."""
    active_entries = await db.timeclock.find(
        {"clock_out": None}, {"_id": 0, "selfie_in": 0, "selfie_out": 0}
    ).to_list(200)
    if not active_entries:
        return {"guards": []}
    uids = list({e["user_id"] for e in active_entries})
    sids = list({e["site_id"] for e in active_entries if e.get("site_id")})
    um = {u["id"]: u for u in await db.users.find({"id": {"$in": uids}}, {"_id": 0, "hashed_password": 0}).to_list(100)}
    sm = {s["id"]: s for s in await db.sites.find({"id": {"$in": sids}}, {"_id": 0}).to_list(50)} if sids else {}
    guards = []
    for entry in active_entries:
        ud = um.get(entry["user_id"])
        if not ud:
            continue
        last_ts = entry.get("last_ping_ts") or entry["clock_in"]
        try:
            stale = (now_utc() - datetime.fromisoformat(last_ts.replace("Z", "+00:00"))).total_seconds() > 900
        except Exception:
            stale = True
        guards.append({
            "user_id": entry["user_id"],
            "full_name": ud.get("full_name"),
            "employee_number": ud.get("employee_number"),
            "photo_url": ud.get("photo_url"),
            "site": sm.get(entry.get("site_id")),
            "clock_in": entry["clock_in"],
            "last_lat": entry.get("last_ping_lat") or entry.get("clock_in_lat"),
            "last_lng": entry.get("last_ping_lng") or entry.get("clock_in_lng"),
            "last_ping_ts": entry.get("last_ping_ts"),
            "stale": stale,
        })
    return {"guards": guards}


# ============================================================
# SHIFT SWAP MARKETPLACE
# ============================================================

@api.post("/shifts/{shift_id}/request-swap", status_code=201)
async def request_swap(shift_id: str, body: SwapRequestIn, user=Depends(get_current_user)):
    shift = await db.shifts.find_one({"id": shift_id, "user_id": user["id"]})
    if not shift:
        raise HTTPException(404, "Shift not found or not yours")
    if shift.get("status") not in ("scheduled",):
        raise HTTPException(400, "Only scheduled shifts can be swapped")
    existing = await db.shift_swaps.find_one({"shift_id": shift_id, "status": "open"})
    if existing:
        raise HTTPException(400, "A swap request for this shift already exists")
    site = None
    if shift.get("site_id"):
        site = await db.sites.find_one({"id": shift["site_id"]}, {"_id": 0, "name": 1, "id": 1})
    swap = {
        "id": str(uuid.uuid4()),
        "shift_id": shift_id,
        "requester_id": user["id"],
        "requester_name": user["full_name"],
        "site_id": shift.get("site_id"),
        "site_name": site["name"] if site else None,
        "start": shift["start"],
        "end": shift["end"],
        "role": shift.get("role"),
        "pay_rate": shift.get("pay_rate"),
        "reason": body.reason,
        "status": "open",
        "volunteer_id": None,
        "volunteer_name": None,
        "created_at": iso(now_utc()),
        "resolved_at": None,
    }
    await db.shift_swaps.insert_one(swap)
    swap.pop("_id", None)
    admin_ids = [u["id"] async for u in db.users.find({"role": "admin"}, {"id": 1, "_id": 0})]
    await send_push(admin_ids, {"title": "Shift Swap Requested", "message": f"{user['full_name']} wants to swap their {shift['start'][:10]} shift."})
    return swap


@api.get("/shift-swaps")
async def list_shift_swaps(user=Depends(get_current_user)):
    own = await db.shift_swaps.find({"requester_id": user["id"]}, {"_id": 0}).sort("created_at", -1).to_list(50)
    marketplace = await db.shift_swaps.find(
        {"status": "open", "requester_id": {"$ne": user["id"]}}, {"_id": 0}
    ).sort("start", 1).to_list(50)
    return {"own": own, "marketplace": marketplace}


@api.post("/shift-swaps/{swap_id}/volunteer")
async def volunteer_swap(swap_id: str, user=Depends(get_current_user)):
    swap = await db.shift_swaps.find_one({"id": swap_id, "status": "open"})
    if not swap:
        raise HTTPException(404, "Swap not found or no longer open")
    if swap["requester_id"] == user["id"]:
        raise HTTPException(400, "Cannot volunteer for your own swap")
    await db.shift_swaps.update_one(
        {"id": swap_id},
        {"$set": {"status": "accepted", "volunteer_id": user["id"], "volunteer_name": user["full_name"]}},
    )
    admin_ids = [u["id"] async for u in db.users.find({"role": "admin"}, {"id": 1, "_id": 0})]
    await send_push(
        admin_ids + [swap["requester_id"]],
        {"title": "Swap Volunteer", "message": f"{user['full_name']} wants to cover {swap['requester_name']}'s shift on {swap['start'][:10]}."},
    )
    return {"volunteered": True}


@api.post("/shift-swaps/{swap_id}/cancel")
async def cancel_swap(swap_id: str, user=Depends(get_current_user)):
    r = await db.shift_swaps.update_one(
        {"id": swap_id, "requester_id": user["id"], "status": {"$in": ["open", "accepted"]}},
        {"$set": {"status": "cancelled", "resolved_at": iso(now_utc())}},
    )
    if r.matched_count == 0:
        raise HTTPException(404, "Swap not found or cannot be cancelled")
    return {"cancelled": True}


@api.post("/admin/shift-swaps/{swap_id}/decision")
async def admin_swap_decision(swap_id: str, body: SwapActionIn, admin=Depends(require_admin)):
    if body.action not in ("approve", "reject"):
        raise HTTPException(400, "action must be 'approve' or 'reject'")
    swap = await db.shift_swaps.find_one({"id": swap_id, "status": "accepted"})
    if not swap:
        raise HTTPException(404, "Swap not found or not in 'accepted' state")
    both_ids = [swap["requester_id"], swap["volunteer_id"]]
    if body.action == "approve":
        await db.shifts.update_one({"id": swap["shift_id"]}, {"$set": {"user_id": swap["volunteer_id"], "swapped_from": swap["requester_id"]}})
        await db.shift_swaps.update_one({"id": swap_id}, {"$set": {"status": "approved", "resolved_at": iso(now_utc()), "approved_by": admin["id"]}})
        await send_push(
            both_ids,
            {"title": "Shift Swap Approved ✓", "message": f"The {swap['start'][:10]} shift has been transferred to {swap['volunteer_name']}."},
        )
        # SMS both parties + in-app
        req_user = await db.users.find_one({"id": swap["requester_id"]},  {"phone": 1, "_id": 0})
        vol_user = await db.users.find_one({"id": swap["volunteer_id"]},  {"phone": 1, "_id": 0})
        await asyncio.gather(
            notif.send_sms(req_user["phone"], f"Skyhawk: Your swap request for {swap['start'][:10]} was APPROVED. You're now off this shift.") if req_user and req_user.get("phone") else asyncio.sleep(0),
            notif.send_sms(vol_user["phone"], f"Skyhawk: Shift swap APPROVED — you're now working {swap['start'][:10]}.") if vol_user and vol_user.get("phone") else asyncio.sleep(0),
            notif.save_inapp(db, both_ids, "Shift Swap Approved ✓", f"The {swap['start'][:10]} shift has been transferred to {swap['volunteer_name']}.", category="swap"),
            return_exceptions=True,
        )
    else:
        await db.shift_swaps.update_one({"id": swap_id}, {"$set": {"status": "rejected", "resolved_at": iso(now_utc()), "rejected_by": admin["id"]}})
        await send_push(
            both_ids,
            {"title": "Shift Swap Rejected", "message": f"The swap request for {swap['start'][:10]} was not approved."},
        )
        await notif.save_inapp(db, both_ids, "Shift Swap Rejected", f"The swap request for {swap['start'][:10]} was not approved.", category="swap")
    return {"action": body.action}


@api.get("/admin/shift-swaps")
async def admin_list_swaps(admin=Depends(require_admin), status: Optional[str] = None):
    q: dict = {}
    if status:
        q["status"] = status
    swaps = await db.shift_swaps.find(q, {"_id": 0}).sort("created_at", -1).to_list(100)
    return {"swaps": swaps}


# ============================================================
# PAYROLL ENGINE
# ============================================================

def _generate_pay_stub_pdf(record: dict) -> str:
    """Render a simple one-page pay stub PDF and return its served URL."""
    filename = f"{record['id']}.pdf"
    path = UPLOAD_DIR / "paystubs" / filename
    c = pdf_canvas.Canvas(str(path), pagesize=letter)
    width, height = letter
    y = height - 72

    c.setFont("Helvetica-Bold", 16)
    c.drawString(72, y, "Skyhawk Security Operations")
    y -= 18
    c.setFont("Helvetica", 10)
    c.drawString(72, y, "Pay Stub")
    y -= 30

    c.setFont("Helvetica-Bold", 11)
    c.drawString(72, y, record["user_name"])
    c.drawString(320, y, f"Employee #: {record.get('employee_number', '-')}")
    y -= 16
    c.setFont("Helvetica", 10)
    c.drawString(72, y, f"Pay period: {record['period_start'][:10]} - {record['period_end'][:10]}")
    y -= 14
    c.drawString(72, y, f"Pay date: {record['pay_date'][:10]}")
    y -= 30

    rows = [
        ("Regular hours", f"{record['hours_regular']:.2f}"),
        ("Overtime hours", f"{record['hours_overtime']:.2f}"),
        ("Hourly rate", f"${record['hourly_rate']:.2f}"),
        ("Overtime multiplier", f"{record['overtime_multiplier']:.2f}x"),
        ("Gross pay", f"${record['gross']:.2f}"),
        ("Tax withheld", f"${record['gross'] - record['net']:.2f}"),
        ("Net pay", f"${record['net']:.2f}"),
    ]
    c.setFont("Helvetica-Bold", 10)
    c.drawString(72, y, "Summary")
    y -= 16
    c.setFont("Helvetica", 10)
    for label, value in rows:
        c.drawString(80, y, label)
        c.drawRightString(300, y, value)
        y -= 14

    y -= 16
    c.setFont("Helvetica-Bold", 10)
    c.drawString(72, y, "Shift detail")
    y -= 16
    c.setFont("Helvetica", 9)
    for item in record.get("line_items", [])[:25]:
        if y < 72:
            c.showPage()
            y = height - 72
        c.drawString(80, y, item.get("date", ""))
        c.drawRightString(300, y, f"{item.get('hours', 0):.2f} h")
        y -= 12

    c.showPage()
    c.save()
    return f"/uploads/paystubs/{filename}"


@api.post("/admin/payroll/calculate", status_code=201)
async def calculate_payroll(body: PayrollCalculateIn, admin=Depends(require_admin)):
    user = await db.users.find_one({"id": body.user_id}, {"_id": 0, "hashed_password": 0})
    if not user:
        raise HTTPException(404, "User not found")
    try:
        ps = datetime.fromisoformat(body.period_start).replace(tzinfo=timezone.utc)
        pe = datetime.fromisoformat(body.period_end).replace(tzinfo=timezone.utc)
        pd_dt = datetime.fromisoformat(body.pay_date).replace(tzinfo=timezone.utc)
    except ValueError:
        raise HTTPException(400, "Invalid date format. Use ISO 8601 (YYYY-MM-DD)")
    entries = await db.timeclock.find(
        {"user_id": body.user_id, "clock_out": {"$ne": None},
         "clock_in": {"$gte": iso(ps), "$lte": iso(pe + timedelta(days=1))}},
        {"_id": 0, "selfie_in": 0, "selfie_out": 0},
    ).to_list(200)
    total_hours = sum(e.get("hours_worked") or 0 for e in entries)
    regular_hours = min(total_hours, body.overtime_threshold)
    overtime_hours = max(0.0, total_hours - body.overtime_threshold)
    gross = round(regular_hours * body.hourly_rate + overtime_hours * body.hourly_rate * body.overtime_multiplier, 2)
    net = round(gross * (1 - body.tax_rate), 2)
    line_items = [{"date": e["clock_in"][:10], "hours": round(e.get("hours_worked") or 0, 2), "shift_id": e.get("shift_id")} for e in entries]
    record = {
        "id": str(uuid.uuid4()),
        "user_id": body.user_id,
        "user_name": user["full_name"],
        "employee_number": user.get("employee_number"),
        "period_start": iso(ps),
        "period_end": iso(pe),
        "pay_date": iso(pd_dt),
        "hours_regular": round(regular_hours, 2),
        "hours_overtime": round(overtime_hours, 2),
        "hourly_rate": body.hourly_rate,
        "overtime_multiplier": body.overtime_multiplier,
        "tax_rate": body.tax_rate,
        "gross": gross,
        "net": net,
        "status": "submitted",
        "line_items": line_items,
        "shifts_count": len(entries),
        "pay_stub_url": None,
        "created_at": iso(now_utc()),
        "created_by": admin["id"],
    }
    record["pay_stub_url"] = _generate_pay_stub_pdf(record)
    await db.payroll.insert_one(record)
    record.pop("_id", None)
    return record


@api.get("/payroll/{period_id}/stub")
async def payroll_stub(period_id: str, user=Depends(get_current_user)):
    q: dict = {"id": period_id}
    if user["role"] != "admin":
        q["user_id"] = user["id"]
    period = await db.payroll.find_one(q, {"_id": 0})
    if not period:
        raise HTTPException(404, "Payroll period not found")
    return period


# ============================================================
# COMPLIANCE / CREDENTIAL MANAGEMENT
# ============================================================

def _compliance_status(expiry_iso: str) -> str:
    try:
        expiry_dt = datetime.fromisoformat(expiry_iso.replace("Z", "+00:00"))
        days_left = (expiry_dt - now_utc()).days
        if days_left < 0:
            return "expired"
        if days_left <= 30:
            return "expiring_soon"
        if days_left <= 60:
            return "expiring"
        return "valid"
    except Exception:
        return "unknown"


@api.get("/compliance/status")
async def my_compliance(user=Depends(get_current_user)):
    docs = await db.wallet_documents.find({"user_id": user["id"]}, {"_id": 0}).to_list(50)
    for doc in docs:
        if doc.get("expiry"):
            doc["compliance_status"] = _compliance_status(doc["expiry"])
            try:
                doc["days_until_expiry"] = (datetime.fromisoformat(doc["expiry"].replace("Z", "+00:00")) - now_utc()).days
            except Exception:
                doc["days_until_expiry"] = None
    expired = [d for d in docs if d.get("compliance_status") == "expired"]
    expiring_soon = [d for d in docs if d.get("compliance_status") == "expiring_soon"]
    expiring = [d for d in docs if d.get("compliance_status") == "expiring"]
    overall = "expired" if expired else ("expiring_soon" if expiring_soon else ("expiring" if expiring else "compliant"))
    return {"overall_status": overall, "documents": docs, "expired_count": len(expired), "expiring_soon_count": len(expiring_soon)}


@api.get("/admin/compliance")
async def admin_compliance(admin=Depends(require_admin)):
    guards = await db.users.find({"role": "employee"}, {"_id": 0, "hashed_password": 0}).to_list(200)
    results = []
    for g in guards:
        docs = await db.wallet_documents.find({"user_id": g["id"]}, {"_id": 0}).to_list(50)
        expired = sum(1 for d in docs if d.get("expiry") and _compliance_status(d["expiry"]) == "expired")
        expiring_soon = sum(1 for d in docs if d.get("expiry") and _compliance_status(d["expiry"]) == "expiring_soon")
        lic_status = _compliance_status(g["licence_expiry"]) if g.get("licence_expiry") else "unknown"
        results.append({
            "user_id": g["id"],
            "full_name": g["full_name"],
            "employee_number": g.get("employee_number"),
            "licence_expiry": g.get("licence_expiry"),
            "licence_status": lic_status,
            "expired_docs": expired,
            "expiring_soon_docs": expiring_soon,
            "total_docs": len(docs),
        })
    results.sort(key=lambda x: (x["licence_status"] not in ("expired",), x["licence_status"] not in ("expiring_soon",)))
    return {"guards": results}


@api.post("/wallet/documents", status_code=201)
async def add_wallet_doc(body: WalletDocIn, user=Depends(get_current_user)):
    doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "type": body.type,
        "name": body.name,
        "number": body.number,
        "expiry": body.expiry,
        "status": _compliance_status(body.expiry),
        "created_at": iso(now_utc()),
    }
    await db.wallet_documents.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.patch("/wallet/documents/{doc_id}")
async def update_wallet_doc(doc_id: str, body: WalletDocIn, user=Depends(get_current_user)):
    existing = await db.wallet_documents.find_one({"id": doc_id, "user_id": user["id"]})
    if not existing:
        raise HTTPException(404, "Document not found")
    update = {"name": body.name, "number": body.number, "expiry": body.expiry, "type": body.type, "status": _compliance_status(body.expiry), "updated_at": iso(now_utc())}
    await db.wallet_documents.update_one({"id": doc_id}, {"$set": update})
    return await db.wallet_documents.find_one({"id": doc_id}, {"_id": 0})


# ============================================================
# COMMUNITY
# ============================================================
def _serialize_post(p: dict, user_id: str) -> dict:
    likes = p.get("likes") or []
    return {
        "id": p["id"],
        "author_id": p.get("author_id"),
        "author_name": p.get("author_name"),
        "author_handle": p.get("author_handle"),
        "author_photo": p.get("author_photo"),
        "audience": p.get("audience", "All Staff"),
        "type": p.get("type", "post"),
        "title": p.get("title"),
        "body": p.get("body", ""),
        "attachments": p.get("attachments") or [],
        "created_at": p.get("created_at"),
        "like_count": len(likes),
        "liked_by_me": user_id in likes,
        "comments": p.get("comments") or [],
        "comment_count": len(p.get("comments") or []),
        "seen_count": len(p.get("seen_by") or []),
    }


@api.get("/community/posts")
async def list_community_posts(type: Optional[str] = Query(default=None), user=Depends(get_current_user)):
    query = {}
    if type and type != "all":
        query["type"] = type
    items = await db.community_posts.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    if items:
        await db.community_posts.update_many(
            {"id": {"$in": [p["id"] for p in items]}},
            {"$addToSet": {"seen_by": user["id"]}},
        )
    return {"posts": [_serialize_post(p, user["id"]) for p in items]}


@api.post("/community/posts", status_code=201)
async def create_community_post(body: CommunityPostIn, user=Depends(get_current_user)):
    if body.type not in ("post", "announcement", "event", "recognition"):
        raise HTTPException(400, "Invalid post type")
    post = {
        "id": str(uuid.uuid4()),
        "author_id": user["id"],
        "author_name": user["full_name"],
        "author_handle": (user["full_name"] or "").lower().replace(" ", ""),
        "author_photo": user.get("photo_url"),
        "audience": body.audience,
        "type": body.type,
        "title": None,
        "body": body.body,
        "attachments": body.attachments,
        "created_at": iso(now_utc()),
        "likes": [],
        "comments": [],
        "seen_by": [user["id"]],
    }
    await db.community_posts.insert_one(post)
    return _serialize_post(post, user["id"])


@api.post("/community/posts/{post_id}/like")
async def toggle_like_community_post(post_id: str, user=Depends(get_current_user)):
    post = await db.community_posts.find_one({"id": post_id}, {"_id": 0})
    if not post:
        raise HTTPException(404, "Post not found")
    likes = post.get("likes") or []
    if user["id"] in likes:
        await db.community_posts.update_one({"id": post_id}, {"$pull": {"likes": user["id"]}})
        liked = False
    else:
        await db.community_posts.update_one({"id": post_id}, {"$addToSet": {"likes": user["id"]}})
        liked = True
    post = await db.community_posts.find_one({"id": post_id}, {"_id": 0})
    return _serialize_post(post, user["id"])


@api.post("/community/posts/{post_id}/comments", status_code=201)
async def add_community_comment(post_id: str, body: CommunityCommentIn, user=Depends(get_current_user)):
    comment = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "user_name": user["full_name"],
        "body": body.body,
        "created_at": iso(now_utc()),
    }
    r = await db.community_posts.update_one({"id": post_id}, {"$push": {"comments": comment}})
    if r.matched_count == 0:
        raise HTTPException(404, "Post not found")
    post = await db.community_posts.find_one({"id": post_id}, {"_id": 0})
    return _serialize_post(post, user["id"])


@api.delete("/admin/community/posts/{post_id}")
async def admin_delete_community_post(post_id: str, admin=Depends(require_admin)):
    r = await db.community_posts.delete_one({"id": post_id})
    if r.deleted_count == 0:
        raise HTTPException(404, "Post not found")
    return {"deleted": True}


@api.delete("/community/posts/{post_id}")
async def delete_own_community_post(post_id: str, user=Depends(get_current_user)):
    q = {"id": post_id}
    if user["role"] != "admin":
        q["author_id"] = user["id"]
    r = await db.community_posts.delete_one(q)
    if r.deleted_count == 0:
        raise HTTPException(404, "Post not found or not authorized")
    return {"deleted": True}


@api.put("/community/posts/{post_id}")
async def edit_own_community_post(post_id: str, body: CommunityEditIn, user=Depends(get_current_user)):
    q = {"id": post_id}
    if user["role"] != "admin":
        q["author_id"] = user["id"]
    r = await db.community_posts.update_one(
        q,
        {"$set": {"body": body.body, "attachments": body.attachments, "updated_at": iso(now_utc())}}
    )
    if r.matched_count == 0:
        raise HTTPException(404, "Post not found or not authorized")
    post = await db.community_posts.find_one({"id": post_id}, {"_id": 0})
    return _serialize_post(post, user["id"])


app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve user-uploaded files (incident/community attachments, generated pay stubs).
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")
